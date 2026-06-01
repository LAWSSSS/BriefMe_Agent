"""xlsx 输出（按用户提供的统计表格图片 1:1 还原）

表头三级合并：
  第1行: 日期 | 车牌 | 工位 | 主料型对比 ........ | 扣杂结果 .........
  第2行:                    | 人工 | 赛迪 | 赛迪准确率% | 赛迪废钢判级差异率统计 | 人工 | 赛迪 | 赛迪扣重扣杂质量差异 | 赛迪扣重占比值
  第3行: —— 料型名 / 占比 / 重量 等实际子字段

经与用户确认的 13 列：
  1 日期  2 车牌  3 工位
  4 人工料型  5 人工%
  6 赛迪料型  7 赛迪%
  8 赛迪准确率%    （按日合并）
  9 赛迪废钢判级差异率统计
  10 人工kg  11 赛迪kg
  12 赛迪扣重扣杂质量差异
  13 赛迪扣重占比值

日期列、赛迪准确率列按日 merge_cells。
主料型不一致行：浅蓝填充。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agent.scrap.dict import get_material_name
from agent.scrap.models import DailyScrapStats, MaterialEntry, TruckStat

logger = logging.getLogger(__name__)


HEADER_LEVEL1 = [
    "日期", "车牌", "工位",
    "主料型对比", "主料型对比", "主料型对比", "主料型对比", "主料型对比", "主料型对比",
    "扣杂结果", "扣杂结果", "扣杂结果", "扣杂结果",
]
HEADER_LEVEL2 = [
    "", "", "",
    "人工", "人工",
    "赛迪", "赛迪",
    "赛迪准确率%",
    "赛迪废钢判级差异率统计",
    "人工",
    "赛迪",
    "赛迪扣重扣杂质量差异",
    "赛迪扣重占比值",
]
HEADER_LEVEL3 = [
    "", "", "",
    "料型", "占比%",
    "料型", "占比%",
    "",
    "",
    "kg", "kg",
    "",
    "",
]

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_MISMATCH = PatternFill("solid", fgColor="BDD7EE")  # 浅蓝
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")  # 淡蓝表头

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "--"
    return f"{v:.2f}%"


def _fmt_num(v: Optional[float], digits: int = 2) -> str:
    if v is None:
        return "--"
    return f"{v:.{digits}f}"


def _material_display(e: Optional[MaterialEntry]) -> str:
    if e is None:
        return "--"
    return get_material_name(e.steel_type, e.steel_level)


def _write_headers(ws, start_row: int = 1) -> int:
    """写入三行表头，返回下一行号"""
    for col_idx, val in enumerate(HEADER_LEVEL1, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=val)
        c.font = Font(bold=True, size=11)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = FILL_HEADER
    for col_idx, val in enumerate(HEADER_LEVEL2, start=1):
        c = ws.cell(row=start_row + 1, column=col_idx, value=val)
        c.font = Font(bold=True, size=11)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = FILL_HEADER
    for col_idx, val in enumerate(HEADER_LEVEL3, start=1):
        c = ws.cell(row=start_row + 2, column=col_idx, value=val)
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = FILL_HEADER

    # 第一级合并：日期(1) 车牌(2) 工位(3) 分别纵向合并 3 行
    for col in (1, 2, 3):
        ws.merge_cells(
            start_row=start_row, start_column=col,
            end_row=start_row + 2, end_column=col,
        )
    # 主料型对比: 4..9 横向合并（第一行）
    ws.merge_cells(
        start_row=start_row, start_column=4,
        end_row=start_row, end_column=9,
    )
    # 扣杂结果: 10..13 横向合并（第一行）
    ws.merge_cells(
        start_row=start_row, start_column=10,
        end_row=start_row, end_column=13,
    )
    # 第二行：
    #   人工: 4..5, 赛迪: 6..7, 赛迪准确率%: 8 (纵向 2 行), 差异率统计: 9 (纵向 2 行)
    #   人工(kg): 10 (纵向 2 行), 赛迪(kg): 11 (纵向 2 行), 扣重差异: 12 (纵向 2 行), 占比值: 13 (纵向 2 行)
    ws.merge_cells(
        start_row=start_row + 1, start_column=4,
        end_row=start_row + 1, end_column=5,
    )
    ws.merge_cells(
        start_row=start_row + 1, start_column=6,
        end_row=start_row + 1, end_column=7,
    )
    for col in (8, 9, 10, 11, 12, 13):
        ws.merge_cells(
            start_row=start_row + 1, start_column=col,
            end_row=start_row + 2, end_column=col,
        )

    return start_row + 3


def _truck_row(stat: TruckStat, accuracy_display: str, mismatch: bool):
    """单车 13 列数据"""
    man = stat.manual_main
    ai = stat.ai_main
    return [
        stat.date,
        stat.car_number,
        stat.station_number,
        _material_display(man),
        _fmt_num(man.rate) if man else "--",
        _material_display(ai),
        _fmt_num(ai.rate) if ai else "--",
        accuracy_display,
        _fmt_pct(stat.diff_rate),
        _fmt_num(stat.manual_deduct_kg),
        _fmt_num(stat.ai_deduct_kg),
        _fmt_num(stat.weight_diff),
        _fmt_num(stat.weight_ratio),
    ]


def write_stats_xlsx(
    stats_list: List[DailyScrapStats],
    save_path: Path,
    title: str = "赛迪废钢判级结果统计",
) -> Path:
    """把多日 stats 写到一个 xlsx 文件

    Returns:
        save_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "废钢检判统计"

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.cell(row=1, column=1).alignment = CENTER

    header_start = 3
    data_start = _write_headers(ws, start_row=header_start)

    row = data_start
    for day in stats_list:
        day_start = row
        acc = day.accuracy_rate
        acc_text = "--" if acc is None else f"{acc:.2f}%"

        for idx, truck in enumerate(day.trucks):
            mismatch = truck.main_same is False
            # 第一辆车显示日期、准确率，其它行留空后面合并
            acc_cell = acc_text if idx == 0 else ""
            date_cell = truck.date if idx == 0 else ""
            row_values = _truck_row(truck, acc_cell, mismatch)
            row_values[0] = date_cell

            for col_idx, val in enumerate(row_values, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.alignment = CENTER
                c.border = BORDER
                if mismatch:
                    c.fill = FILL_MISMATCH
            row += 1

        # 合并日期列和准确率列
        if len(day.trucks) > 1:
            ws.merge_cells(
                start_row=day_start, start_column=1,
                end_row=row - 1, end_column=1,
            )
            ws.merge_cells(
                start_row=day_start, start_column=8,
                end_row=row - 1, end_column=8,
            )
        if len(day.trucks) >= 1:
            ws.cell(row=day_start, column=1, value=day.date).alignment = CENTER
            ws.cell(row=day_start, column=8, value=acc_text).alignment = CENTER

    widths = [12, 11, 6, 12, 9, 12, 9, 12, 16, 10, 10, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    logger.info("xlsx 已保存: %s", save_path)
    return save_path


WEEKLY_TYPE_HEADERS = [
    "数据源", "开始日期", "统计日期", "主料型",
    "车数", "主料型准确数量",
    "主料型占比差异（<10%）数量", "扣重误差+-100KG数量",
    "主料型准确率", "主料型占比准确率（误差<=10%）",
    "平均扣重误差KG（误差<=100KG）", "扣杂误差+-100KG准确率",
    "扣杂误差占比0.5~1.5",
]


def _norm_pct(v: float) -> float:
    if abs(v) < 1.0 and v != 0.0:
        return v * 100.0
    return v


def _fmt_pct_or_dash(v: Optional[float]) -> str:
    if v is None:
        return "-"
    return f"{_norm_pct(v):.2f}%"


def _fmt_num_or_dash(v: Optional[float], digits: int = 2) -> str:
    if v is None:
        return "-"
    formatted = f"{v:.{digits}f}"
    if "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def _fmt_date_for_excel(date_str: str) -> str:
    parts = date_str.split("-")
    return f"{parts[0]}/{int(parts[1])}/{int(parts[2])}"


def _write_weekly_summary(
    ws, row: int, label: str,
    truck_count: int,
    main_same_count: int,
    ratio_10pct_count: int,
    weight_100kg_count: int,
    main_same_pct: Optional[float],
    ratio_10pct_pct: Optional[float],
    avg_weight_diff_kg: Optional[float],
    weight_100kg_pct: Optional[float],
    deduct_ratio: Optional[float],
) -> None:
    vals = [
        "",
        "",
        "",
        label,
        truck_count if truck_count > 0 else "",
        main_same_count,
        ratio_10pct_count,
        weight_100kg_count,
        _fmt_pct_or_dash(main_same_pct),
        _fmt_pct_or_dash(ratio_10pct_pct),
        _fmt_num_or_dash(avg_weight_diff_kg),
        _fmt_pct_or_dash(weight_100kg_pct),
        _fmt_num_or_dash(deduct_ratio),
    ]
    for col_idx, val in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.alignment = CENTER
        c.border = BORDER
        c.font = Font(bold=True)


def write_weekly_type_xlsx(
    report,
    save_path: Path,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "周度料型统计"

    start_display = _fmt_date_for_excel(report.start_date)
    end_display = _fmt_date_for_excel(report.end_date)

    title = f"镔鑫料型统计_{report.start_date}_{report.end_date}"
    ncols = len(WEEKLY_TYPE_HEADERS)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).alignment = CENTER

    header_row = 3
    for col_idx, val in enumerate(WEEKLY_TYPE_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=val)
        c.font = Font(bold=True, size=11)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = FILL_HEADER

    row = header_row + 1
    first_data_row = row

    for idx, r in enumerate(report.rows):
        has_eligible = r.main_accuracy_pct is not None
        is_first = (idx == 0)
        vals = [
            "赛迪" if is_first else "",
            start_display if is_first else "",
            end_display if is_first else "",
            r.material_name,
            r.truck_count if r.truck_count > 0 else "",
            r.main_correct_count if has_eligible else "",
            r.ratio_within_10pct_count if has_eligible else "",
            r.weight_within_100kg_count if has_eligible else "",
            _fmt_pct_or_dash(r.main_accuracy_pct) if has_eligible else "",
            _fmt_pct_or_dash(r.ratio_accuracy_pct) if has_eligible else "",
            _fmt_num_or_dash(r.avg_weight_diff_kg) if has_eligible else "",
            _fmt_pct_or_dash(r.weight_accuracy_pct) if has_eligible else "",
            _fmt_num_or_dash(r.deduct_ratio) if has_eligible else "",
        ]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = CENTER
            c.border = BORDER
        row += 1

    last_data_row = row - 1
    if last_data_row > first_data_row:
        for col in (1, 2, 3):
            ws.merge_cells(
                start_row=first_data_row, start_column=col,
                end_row=last_data_row, end_column=col,
            )

    _write_weekly_summary(ws, row, "整体",
        report.overall_truck_count,
        report.overall_main_same_count,
        report.overall_ratio_within_10pct_count,
        report.overall_weight_within_100kg_count,
        report.overall_main_same_pct,
        report.overall_ratio_within_10pct_pct,
        report.overall_avg_weight_diff_kg,
        report.overall_weight_within_100kg_count / report.overall_truck_count * 100.0
        if report.overall_truck_count > 0 else None,
        report.overall_deduct_ratio,
    )
    row += 1

    _write_weekly_summary(ws, row, "整体（不含中废、杂模）",
        report.no_exclude_truck_count,
        report.no_exclude_main_same_count,
        report.no_exclude_ratio_within_10pct_count,
        report.no_exclude_weight_within_100kg_count,
        report.no_exclude_main_same_pct,
        report.no_exclude_ratio_within_10pct_pct,
        report.no_exclude_avg_weight_diff_kg,
        report.no_exclude_weight_within_100kg_count / report.no_exclude_truck_count * 100.0
        if report.no_exclude_truck_count > 0 else None,
        report.no_exclude_deduct_ratio,
    )

    max_material_len = max(
        (len(r.material_name) for r in report.rows), default=4
    )
    col4_width = max(14, max_material_len * 2 + 2)

    widths = [12, 12, 12, col4_width, 8, 14, 18, 16, 14, 18, 18, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(first_data_row, row + 1):
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = ws.cell(row=first_data_row, column=1)

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    logger.info("周度料型 xlsx 已保存: %s", save_path)
    return save_path
