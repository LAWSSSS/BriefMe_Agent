"""盛隆废钢 xlsx 输出（41列重排版：含最终结算单价、人工详情展开与累计统计）

41 列总布局（与图 3 对齐并扩展单价列）：
  A-C   日期 / 车牌 / 工位
  D-I   主料型对比：人工料型|人工占比%|AI料型|AI占比%|正确?|差异值(%)
  J-M   扣重对比(吨)：人工|AI|比值(K/J)|误差(K-J)
  N-Q   单价对比(元)：人工|AI|最终结算单价|差异
  R-Z   人工结果详情①主料型结果(%)：3 人 × (姓名|料型|占比%) = 9 列
  AA-AF 人工结果详情②扣重结果(吨)：3 人 × (姓名|数值) = 6 列
  AG-AO 人工结果详情③单价(元)：3 人 × (姓名|网页单价|计算单价) = 9 列

表头 3 行：大类 / 中类（人工详情区在这里合并了 2 行高度） / 子类
主料不一致行浅红底色；
末尾加识别率 R 和扣杂符合率汇总框（红色边框）；
**人工姓名按每辆车实际 manual_operators 填，不同车不同名。**

两种入口：
  · ``write_stats_xlsx``      单周期（统计周期概括 + 累计统计 + 详情）
  · ``write_master_xlsx``     多周期主表（Sheet1 多个 14 行块 + 累计统计 + 详情多段，
                              环比自动链：第 i 周期 prev = 第 i-1 周期实际值）
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agent.shenglong.dict import get_material_name, STEEL_TYPE_PRICE
from agent.shenglong.models import (
    DailyShenglongStats,
    ManualOperator,
    MaterialRate,
    PeriodSummary,
    TruckStat,
)

logger = logging.getLogger(__name__)

TOTAL_COLS = 41

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED_SIDE = Side(border_style="medium", color="C00000")
RED_BORDER = Border(left=RED_SIDE, right=RED_SIDE, top=RED_SIDE, bottom=RED_SIDE)

FILL_HEADER_L1 = PatternFill("solid", fgColor="D9E1F2")
FILL_HEADER_L2 = PatternFill("solid", fgColor="E7EEF9")
FILL_HEADER_L3 = PatternFill("solid", fgColor="F2F5FB")
FILL_MISMATCH = PatternFill("solid", fgColor="FCE4E4")  # 浅红
FILL_SUMMARY = PatternFill("solid", fgColor="FFF2CC")  # 浅黄汇总
FILL_CUM_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_CUM_GROUP = PatternFill("solid", fgColor="5B9BD5")
FILL_CUM_HEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_CUM_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_CUM_GOOD = PatternFill("solid", fgColor="E2F0D9")
FILL_CUM_BAD = PatternFill("solid", fgColor="FCE4D6")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "/"
    return f"{v:.2f}%"


def _fmt_num(v: Optional[float], digits: int = 2) -> str:
    if v is None:
        return "/"
    return f"{v:.{digits}f}"


def _fmt_bool_yn(v: Optional[bool]) -> str:
    if v is None:
        return ""
    return "是" if v else "否"


def _rate_fill(
    numerator: int,
    denominator: int,
    target_rate: float,
) -> PatternFill:
    if denominator <= 0:
        return FILL_CUM_HEADER
    return FILL_CUM_GOOD if numerator / denominator >= target_rate else FILL_CUM_BAD


def _material_display(m: Optional[MaterialRate]) -> str:
    if m is None:
        return "--"
    return get_material_name(m.steel_type)


def _get_operator_at(truck: TruckStat, idx: int) -> Optional[ManualOperator]:
    """取这辆车 manual_operators 的第 idx 个（0-based），不足返回 None。

    不做跨车匹配——每辆车各自的 3 人都可能不同，不应以"第一辆车的姓名"去找"第二辆车的同名人"。
    """
    if 0 <= idx < len(truck.manual_operators):
        return truck.manual_operators[idx]
    return None


def _op_calc_price(op: Optional[ManualOperator]) -> Optional[float]:
    """计算单个检验员的加权单价"""
    if op is None:
        return None
    total = 0.0
    has_any = False
    for m in op.materials:
        if m.steel_type is None:
            continue
        unit = STEEL_TYPE_PRICE.get(int(m.steel_type))
        if unit is None:
            continue
        total += (m.rate / 100.0) * unit
        has_any = True
    return total if has_any else None


# ======================================================================
#  表头写入
# ======================================================================
def _fill_header_row(
    ws, row: int, values: List[str], fill: PatternFill, font_size: int = 11
):
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(bold=True, size=font_size)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = fill


def _write_headers(ws, start_row: int) -> int:
    """写三行表头，返回数据起始行号。"""

    # --- 第 1 行：大类 ---
    row1 = [""] * TOTAL_COLS
    row1[0], row1[1], row1[2] = "日期", "车牌", "工位"
    # D-I (4..9) 主料型对比
    for i in range(3, 9):
        row1[i] = "主料型对比"
    # J-M (10..13) 扣重对比(吨)
    for i in range(9, 13):
        row1[i] = "扣重对比(吨)"
    # N-Q (14..17) 单价对比(元)
    for i in range(13, 17):
        row1[i] = "单价对比(元)"
    # R-Z (18..26) 人工结果详情①
    for i in range(17, 26):
        row1[i] = "人工结果详情①"
    # AA-AF (27..32) 人工结果详情②
    for i in range(26, 32):
        row1[i] = "人工结果详情②"
    # AG-AO (33..41) 人工结果详情③
    for i in range(32, 41):
        row1[i] = "人工结果详情③"
    _fill_header_row(ws, start_row, row1, FILL_HEADER_L1)

    # --- 第 2 行：中类 ---
    row2 = [""] * TOTAL_COLS
    # D-E: 人工；F-G: AI；H: 正确?；I: 差异值(%)
    row2[3], row2[5], row2[7], row2[8] = "人工", "AI", "正确?", "差异值(%)"
    # J: 人工；K: AI；L: 比值(K/J)；M: 误差(K-J)
    row2[9], row2[10], row2[11], row2[12] = "人工", "AI", "比值(K/J)", "误差(K-J)"
    # N: 人工；O: AI；P: 最终结算单价；Q: 差异
    row2[13], row2[14], row2[15], row2[16] = "人工", "AI", "最终结算单价", "差异"
    
    # 详情大块横向合并（第2行）
    row2[17] = "主料型结果（%）"
    row2[26] = "扣重结果（吨）"
    row2[32] = "单价结果"
    _fill_header_row(ws, start_row + 1, row2, FILL_HEADER_L2)

    # --- 第 3 行：子类 ---
    row3 = [""] * TOTAL_COLS
    # D-G 主料型对比子表头
    row3[3], row3[4], row3[5], row3[6] = "料型", "占比%", "料型", "占比%"
    
    # R-Z 人工详情①
    for i in range(3):
        base = 17 + i * 3
        row3[base], row3[base + 1], row3[base + 2] = "姓名", "料型", "占比%"
    
    # AA-AF 人工详情②
    for i in range(3):
        base = 26 + i * 2
        row3[base], row3[base + 1] = "姓名", "数值"
        
    # AG-AO 人工详情③
    for i in range(3):
        base = 32 + i * 3
        row3[base], row3[base + 1], row3[base + 2] = "姓名", "网页单价", "计算单价"
        
    _fill_header_row(ws, start_row + 2, row3, FILL_HEADER_L3, font_size=10)

    # --- 合并 ---
    # A/B/C 纵合并 3 行
    for col in (1, 2, 3):
        ws.merge_cells(
            start_row=start_row, start_column=col, end_row=start_row + 2, end_column=col
        )

    # 第 1 行大类合并：D-I, J-M, N-Q, R-Z, AA-AF, AG-AO
    _h1_spans = [
        (4, 9), (10, 13), (14, 17), (18, 26), (27, 32), (33, 41)
    ]
    for s, e in _h1_spans:
        ws.merge_cells(
            start_row=start_row, start_column=s, end_row=start_row, end_column=e
        )

    # 第 2 行横合并：人工、AI 以及 详情①②③ 的大块
    _h2_h_spans = [(4, 5), (6, 7), (18, 26), (27, 32), (33, 41)]
    for s, e in _h2_h_spans:
        ws.merge_cells(
            start_row=start_row + 1, start_column=s, end_row=start_row + 1, end_column=e
        )

    # 第 2~3 行纵合并：H(8), I(9), J-M(10..13), N-Q(14..17)
    _h2_v_cols = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
    for col in _h2_v_cols:
        ws.merge_cells(
            start_row=start_row + 1, start_column=col, end_row=start_row + 2, end_column=col
        )

    return start_row + 3


# ======================================================================
#  数据行
# ======================================================================
def _truck_row_values(truck: TruckStat) -> List[object]:
    """按 41 列顺序返回一辆车的值。人工详情区按该车自己的 3 人排列。"""
    man = truck.manual_main
    ai = truck.ai_main
    row: List[object] = [""] * TOTAL_COLS

    row[0] = truck.date  # A
    row[1] = truck.car_number  # B
    row[2] = truck.station_number  # C

    # D-I 主料型对比
    row[3] = _material_display(man)
    row[4] = _fmt_num(man.rate) if man else "/"
    row[5] = _material_display(ai)
    row[6] = _fmt_num(ai.rate) if ai else "/"
    row[7] = _fmt_bool_yn(truck.main_same)
    row[8] = _fmt_num(truck.diff_rate)

    # J-M 扣重对比(吨)
    row[9] = _fmt_num(truck.manual_deduct_ton, digits=3)
    row[10] = _fmt_num(truck.ai_deduct_ton, digits=3)
    row[11] = _fmt_num(truck.weight_ratio, digits=3)
    row[12] = _fmt_num(truck.weight_diff_ton, digits=3)

    # N-Q 单价对比(元)
    row[13] = _fmt_num(truck.manual_steel_price, digits=2)
    row[14] = _fmt_num(truck.ai_steel_price, digits=2)
    row[15] = _fmt_num(truck.final_steel_price, digits=2)
    row[16] = _fmt_num(truck.price_diff, digits=2)

    # R-Z 人工详情① 主料(%)：3 人 × (姓名, 料型, 占比)
    anchor_type = (
        truck.manual_main.steel_type if truck.manual_main is not None else None
    )
    anchor_name = (
        get_material_name(anchor_type) if anchor_type is not None else ""
    )
    for idx in range(3):
        op = _get_operator_at(truck, idx)
        base = 17 + idx * 3
        if op is None:
            row[base], row[base + 1], row[base + 2] = "", "", ""
            continue
        row[base] = op.name or ""
        if anchor_type is None:
            row[base + 1], row[base + 2] = "", ""
            continue
        op_rate = next(
            (m.rate for m in op.materials if m.steel_type == anchor_type),
            None,
        )
        if op_rate is None:
            row[base + 1], row[base + 2] = "", ""
        else:
            row[base + 1] = anchor_name
            row[base + 2] = _fmt_num(op_rate)

    # AA-AF 人工详情② 扣重(吨)：3 人 × (姓名, 数值)
    for idx in range(3):
        op = _get_operator_at(truck, idx)
        base = 26 + idx * 2
        if op is None:
            row[base], row[base + 1] = "", ""
        else:
            row[base] = op.name or ""
            row[base + 1] = _fmt_num(op.deduction_ton, digits=3)

    # AG-AO 人工详情③ 单价(元)：3 人 × (姓名, 网页单价, 计算单价)
    for idx in range(3):
        op = _get_operator_at(truck, idx)
        base = 32 + idx * 3
        if op is None:
            row[base], row[base + 1], row[base + 2] = "", "", ""
        else:
            row[base] = op.name or ""
            row[base + 1] = _fmt_num(op.steel_price, digits=2)
            row[base + 2] = _fmt_num(_op_calc_price(op), digits=2)

    return row


def _apply_row_style(ws, row_idx: int, mismatch: bool) -> None:
    for col in range(1, TOTAL_COLS + 1):
        c = ws.cell(row=row_idx, column=col)
        c.alignment = CENTER
        c.border = BORDER
        if mismatch:
            c.fill = FILL_MISMATCH


# ======================================================================
#  主入口
# ======================================================================
def write_stats_xlsx(
    stats_list: List[DailyShenglongStats],
    save_path: Path,
    title: str = "盛隆赛迪废钢判级结果统计",
    target_recognition_rate: float = 0.92,
    target_deduction_compliance_rate: float = 0.90,
    period_summary: Optional[PeriodSummary] = None,
) -> Path:
    """把多日 stats 写到一个 xlsx 文件：
        Sheet1「统计周期概括」 —— period_summary（可选；缺省由 stats_list 推算）
        Sheet2「累计统计」 —— 截图式累计汇总（period_summary 存在时）
        Sheet3「检判统计详情」 —— 41 列详情表
    """
    wb = Workbook()

    # ============ Sheet1 统计周期概括 ============
    if period_summary is not None:
        ws_summary = wb.active
        ws_summary.title = "统计周期概括"
        _write_summary_sheet(ws_summary, [period_summary])
        ws_cumulative = wb.create_sheet(title="累计统计")
        _write_cumulative_sheet(
            ws_cumulative,
            [period_summary],
            target_recognition_rate=target_recognition_rate,
            target_deduction_compliance_rate=target_deduction_compliance_rate,
        )
        ws = wb.create_sheet(title="检判统计详情")
    else:
        ws = wb.active
        ws.title = "检判统计详情"

    # ============ 检判统计详情（41 列逻辑） ============
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    ws.cell(row=1, column=1).alignment = CENTER

    header_start = 3
    data_start = _write_headers(ws, header_start)

    row = data_start
    for day in stats_list:
        if not day.trucks:
            continue
        day_start = row
        for idx, truck in enumerate(day.trucks):
            vals = _truck_row_values(truck)
            if idx != 0:
                vals[0] = ""
            for col_idx, v in enumerate(vals, start=1):
                ws.cell(row=row, column=col_idx, value=v)
            mismatch = truck.main_same is False
            _apply_row_style(ws, row, mismatch)
            row += 1

        if len(day.trucks) > 1:
            ws.merge_cells(
                start_row=day_start, start_column=1,
                end_row=row - 1, end_column=1,
            )
        ws.cell(row=day_start, column=1, value=day.date).alignment = CENTER

    if row > data_start:
        row += 1
        _write_summary_box(
            ws,
            row,
            stats_list,
            target_recognition_rate,
            target_deduction_compliance_rate,
        )

    _apply_sheet2_widths(ws)

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    logger.info("xlsx 已保存: %s", save_path)
    return save_path


# ======================================================================
#  「检判统计详情」列宽（多入口共享）
# ======================================================================
def _apply_sheet2_widths(ws) -> None:
    """统一设定「检判统计详情」41 列的列宽。"""
    widths = [
        12, 11, 6,             # A-C 日期/车牌/工位 (3)
        10, 8, 10, 8, 7, 10,   # D-I 主料型对比 (6)
        10, 10, 10, 10,        # J-M 扣重对比 (4)
        10, 10, 10, 10,        # N-Q 单价对比 (4)
        9, 8, 6,               # R-T 人工1：姓名|料型|占比 (3)
        9, 8, 6,               # U-W 人工2：姓名|料型|占比 (3)
        9, 8, 6,               # X-Z 人工3：姓名|料型|占比 (3)
        9, 8,                  # AA-AB 人工1：姓名|扣重 (2)
        9, 8,                  # AC-AD 人工2：姓名|扣重 (2)
        9, 8,                  # AE-AF 人工3：姓名|扣重 (2)
        9, 8, 6,               # AG-AI 人工1：姓名|网页单价|计算单价 (3)
        9, 8, 6,               # AJ-AL 人工2：姓名|网页单价|计算单价 (3)
        9, 8, 6,               # AM-AO 人工3：姓名|网页单价|计算单价 (3)
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ======================================================================
#  Sheet2 「写一个周期段」 — 用于多周期主表
# ======================================================================
SECTION_TITLE_FILL = PatternFill("solid", fgColor="305496")  # 周期段标题深蓝
SECTION_TITLE_FONT = Font(bold=True, size=13, color="FFFFFF")


def _write_period_section(
    ws,
    start_row: int,
    stats_list: List[DailyShenglongStats],
    period_idx: int,
    period_label: str,
    target_r: float,
    target_c: float,
) -> int:
    """在 Sheet2 的 ``start_row`` 起，写一个周期的完整段：
        · 周期标题（深蓝跨 41 列）
        · 三级表头（3 行）
        · 数据（按日期分组）
        · 期间汇总框（5 行）
        · 末尾空 1 行
    返回下一段可用的行号。
    """
    r = start_row

    # ---- 周期标题 ----
    cell = ws.cell(row=r, column=1, value=period_label)
    cell.font = SECTION_TITLE_FONT
    cell.alignment = CENTER
    ws.merge_cells(
        start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS
    )
    for col in range(1, TOTAL_COLS + 1):
        ws.cell(row=r, column=col).fill = SECTION_TITLE_FILL
    ws.row_dimensions[r].height = 22
    r += 1

    # ---- 三级表头 ----
    data_start = _write_headers(ws, r)
    r = data_start

    # ---- 数据 ----
    for day in stats_list:
        if not day.trucks:
            continue
        day_start = r
        for idx, truck in enumerate(day.trucks):
            vals = _truck_row_values(truck)
            if idx != 0:
                vals[0] = ""
            for col_idx, v in enumerate(vals, start=1):
                ws.cell(row=r, column=col_idx, value=v)
            mismatch = truck.main_same is False
            _apply_row_style(ws, r, mismatch)
            r += 1
        if len(day.trucks) > 1:
            ws.merge_cells(
                start_row=day_start, start_column=1,
                end_row=r - 1, end_column=1,
            )
        ws.cell(row=day_start, column=1, value=day.date).alignment = CENTER

    # ---- 期间汇总（5 行：标题 + 4 条统计）----
    if r > data_start:
        r += 1
        r = _write_summary_box(ws, r, stats_list, target_r, target_c)

    # ---- 段间空 1 行 ----
    r += 1
    return r


# ======================================================================
#  多周期主表入口
# ======================================================================
def write_master_xlsx(
    cycles: List[Tuple[List[DailyShenglongStats], PeriodSummary]],
    save_path: Path,
    *,
    title: str = "盛隆赛迪废钢判级结果统计 · 主表",
    target_recognition_rate: float = 0.92,
    target_deduction_compliance_rate: float = 0.90,
    auto_link_prev: bool = True,
) -> Path:
    """生成多周期主表：

    Sheet1「统计周期概括」会把每个周期的 14 行块依次往下排（A 列 1/2/3...）。
    Sheet2「累计统计」汇总各期和 Tol 合计。
    Sheet3「检判统计详情」每个周期一段（深蓝段标题 + 三级表头 + 数据 + 期间汇总）。

    Args:
        cycles: 周期列表 [(stats_list, period_summary), ...]，**按时间升序**
        save_path: 输出路径
        auto_link_prev: 自动建立环比链——第 i 周期的 prev_* 设为第 i-1 周期的
            实际识别率 / 扣重符合率（用户已经手设的非 None 值会优先保留）

    Returns:
        save_path
    """
    if not cycles:
        raise ValueError("cycles 不能为空")

    # ---- 自动环比链 ----
    if auto_link_prev:
        for i in range(1, len(cycles)):
            _, p_cur = cycles[i]
            _, p_prev = cycles[i - 1]
            if p_cur.prev_recognition_rate is None and p_prev.recognition_rate_pct is not None:
                p_cur.prev_recognition_rate = p_prev.recognition_rate_pct / 100.0
            if (
                p_cur.prev_deduction_compliance_rate is None
                and p_prev.deduction_compliance_rate_pct is not None
            ):
                p_cur.prev_deduction_compliance_rate = (
                    p_prev.deduction_compliance_rate_pct / 100.0
                )
            if p_cur.prev_cycle_label is None:
                p_cur.prev_cycle_label = p_prev.cycle_label

    # ---- 注入累计指标（首期累计到当期）----
    cum_main_within = 0
    cum_judgable = 0
    cum_dd_ok = 0
    for _, p in cycles:
        cum_main_within += p.main_within_10pct_count
        cum_judgable += p.judgable_trucks
        cum_dd_ok += p.deduction_compliant_count
        p.cumulative_recognition_rate = (
            cum_main_within / cum_judgable if cum_judgable > 0 else None
        )
        p.cumulative_deduction_compliance_rate = (
            cum_dd_ok / cum_judgable if cum_judgable > 0 else None
        )

    # ---- 构建工作簿 ----
    wb = Workbook()

    # ============ Sheet1 统计周期概括（多 14 行块）============
    ws_summary = wb.active
    ws_summary.title = "统计周期概括"
    _write_summary_sheet(ws_summary, [p for _, p in cycles])

    # ============ Sheet2 累计统计（用户截图式总览）============
    ws_cumulative = wb.create_sheet(title="累计统计")
    _write_cumulative_sheet(
        ws_cumulative,
        [p for _, p in cycles],
        target_recognition_rate=target_recognition_rate,
        target_deduction_compliance_rate=target_deduction_compliance_rate,
    )

    # ============ Sheet3 检判统计详情（多段）============
    ws = wb.create_sheet(title="检判统计详情")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 24

    r = 3
    for idx, (stats_list, period) in enumerate(cycles, start=1):
        car_count = sum(len(d.trucks) for d in stats_list)
        period_label = (
            f"第 {idx} 期 · {period.cycle_label} （共 {car_count} 车）"
        )
        r = _write_period_section(
            ws, r, stats_list, idx, period_label,
            target_recognition_rate, target_deduction_compliance_rate,
        )

    _apply_sheet2_widths(ws)

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    logger.info(
        "主表 xlsx 已保存（%d 个周期）: %s", len(cycles), save_path
    )
    return save_path


# =====================================================================
#  Sheet1 「统计周期概括」 —— 14 行/周期模板（参照参考表）
# =====================================================================
PERIOD_BLOCK_ROWS = 14
SUMMARY_TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")  # 段落标题（识别率/扣重符合率/价格差异）
SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")  # 表头（条件/结果/目标值）
SUMMARY_RESULT_FILL = PatternFill("solid", fgColor="FFF2CC")  # 关键结果格底色


def _set_summary_cell(
    ws,
    row: int,
    col: int,
    value,
    *,
    bold: bool = False,
    fill: Optional[PatternFill] = None,
    wrap: bool = True,
    number_format: Optional[str] = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.font = Font(bold=bold, size=11)
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if number_format is not None:
        cell.number_format = number_format


def _merge_with_style(ws, start_row, start_col, end_row, end_col) -> None:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col,
    )


def _write_one_period_block(
    ws, base_row: int, cycle_idx: int, p: PeriodSummary
) -> None:
    """写一个周期的 14 行块，base_row 是该块第一行的 1-based 行号

    F/G 列（"指标变化（环比）"）只有在 ``p.prev_recognition_rate`` /
    ``p.prev_deduction_compliance_rate`` 为非 None 时才填实际值；
    否则 F6/F10 留空、G 列环比也写"/"，避免误导。
    """
    r = base_row

    # ---- A 列序号（合并 14 行）----
    _set_summary_cell(ws, r, 1, cycle_idx, bold=True, fill=SUMMARY_HEADER_FILL)
    _merge_with_style(ws, r, 1, r + PERIOD_BLOCK_ROWS - 1, 1)
    # 合并后给所有行的 A 列都加边框（openpyxl 合并单元格只保留左上角样式，内部要手动补）
    for rr in range(r + 1, r + PERIOD_BLOCK_ROWS):
        ws.cell(row=rr, column=1).border = BORDER

    # ---- F/G/H 列：指标变化大标题（合并 F1:H3 区域，纵跨第 1~3 行）----
    _set_summary_cell(ws, r, 6, "指标变化", bold=True, fill=SUMMARY_TITLE_FILL)
    _merge_with_style(ws, r, 6, r + 2, 8)
    # 合并区内部其他单元格也补边框/底色
    for rr in range(r, r + 3):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).border = BORDER
            if rr == r and cc == 6:
                continue
            ws.cell(row=rr, column=cc).fill = SUMMARY_TITLE_FILL

    # ---- 行 1：统计周期 ----
    _set_summary_cell(ws, r, 2, "统计周期", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r, 3, p.cycle_label, bold=True)
    _merge_with_style(ws, r, 3, r, 5)
    for cc in (4, 5):
        ws.cell(row=r, column=cc).border = BORDER

    # ---- 行 2：周期内有效检判车次数 ----
    _set_summary_cell(ws, r + 1, 2, "周期内有效检判车次数", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 1, 3, p.judgable_trucks, bold=True)
    _merge_with_style(ws, r + 1, 3, r + 1, 5)
    for cc in (4, 5):
        ws.cell(row=r + 1, column=cc).border = BORDER

    # ---- 行 3：识别率 段落标题 ----
    _set_summary_cell(
        ws, r + 2, 2, p.recognition_section_title,
        bold=True, fill=SUMMARY_TITLE_FILL,
    )
    _merge_with_style(ws, r + 2, 2, r + 2, 5)
    for cc in (3, 4, 5):
        ws.cell(row=r + 2, column=cc).border = BORDER
        ws.cell(row=r + 2, column=cc).fill = SUMMARY_TITLE_FILL

    # ---- 行 4：识别率 表头（含 F/G/H 上周期/环比/累计小表头）----
    _set_summary_cell(ws, r + 3, 2, "条件1", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 3, 3, "条件2", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 3, 4, "结果", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 3, 5, "目标值", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 3, 6, "上周期结果", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 3, 7, "环比", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(
        ws, r + 3, 8, p.cumulative_recognition_label,
        bold=True, fill=SUMMARY_HEADER_FILL,
    )

    # ---- 行 5：识别率 子表头（含 F5 "识别准确率"）----
    _set_summary_cell(
        ws, r + 4, 2, p.recognition_condition1_label,
        fill=SUMMARY_HEADER_FILL,
    )
    _set_summary_cell(
        ws, r + 4, 3, p.recognition_condition2_label,
        fill=SUMMARY_HEADER_FILL,
    )
    _set_summary_cell(
        ws, r + 4, 4, p.recognition_result_label,
        fill=SUMMARY_HEADER_FILL,
    )
    _set_summary_cell(
        ws,
        r + 4,
        5,
        "第一阶段 ≥70%\n第二阶段 ≥80%\n第三阶段 ≥92%",
        fill=SUMMARY_HEADER_FILL,
    )
    _merge_with_style(ws, r + 4, 5, r + 5, 5)
    ws.cell(row=r + 5, column=5).border = BORDER
    # F5 = 识别率标签
    _set_summary_cell(
        ws, r + 4, 6, p.recognition_result_label,
        fill=SUMMARY_HEADER_FILL,
    )
    # G5:G6 合并 = 环比公式（写在 G5/G6 中的左上角，整段合并跨 r+4 ~ r+5）
    # H5:H6 合并 = 累计准确率值（小数 0~1）
    if p.cumulative_recognition_rate is not None:
        _set_summary_cell(
            ws, r + 4, 8, float(p.cumulative_recognition_rate), bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
    else:
        _set_summary_cell(ws, r + 4, 8, "", fill=SUMMARY_HEADER_FILL)
    _merge_with_style(ws, r + 4, 8, r + 5, 8)
    ws.cell(row=r + 5, column=8).border = BORDER
    if p.cumulative_recognition_rate is None:
        ws.cell(row=r + 5, column=8).fill = SUMMARY_HEADER_FILL

    # ---- 行 6：识别率 数值 + F6 上周期数值 + G5:G6 环比 ----
    _set_summary_cell(ws, r + 5, 2, p.main_name_match_count, bold=True)
    _set_summary_cell(ws, r + 5, 3, p.main_within_10pct_count, bold=True)
    formula = f"=IFERROR(C{r + 5}/C{r + 1},0)"
    _set_summary_cell(
        ws, r + 5, 4, formula, bold=True, fill=SUMMARY_RESULT_FILL,
        number_format="0.00%",
    )
    # F6 上周期识别率（小数 0~1）
    if p.prev_recognition_rate is not None:
        _set_summary_cell(
            ws, r + 5, 6, float(p.prev_recognition_rate), bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
        # G5:G6 合并 + 环比公式 =(D6-F6)/F6
        ratio_formula = f"=IFERROR((D{r + 5}-F{r + 5})/F{r + 5},0)"
        _set_summary_cell(
            ws, r + 4, 7, ratio_formula, bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
        _merge_with_style(ws, r + 4, 7, r + 5, 7)
        ws.cell(row=r + 5, column=7).border = BORDER
    else:
        # 首期：F6 留空、G 写 "/"，且 G5:G6 合并
        _set_summary_cell(ws, r + 5, 6, "", fill=SUMMARY_HEADER_FILL)
        _set_summary_cell(ws, r + 4, 7, "/", bold=True, fill=SUMMARY_HEADER_FILL)
        _merge_with_style(ws, r + 4, 7, r + 5, 7)
        ws.cell(row=r + 5, column=7).border = BORDER

    # ---- 行 7：扣重符合率 段落标题 ----
    _set_summary_cell(ws, r + 6, 2, "扣重符合率", bold=True, fill=SUMMARY_TITLE_FILL)
    _merge_with_style(ws, r + 6, 2, r + 6, 5)
    for cc in (3, 4, 5):
        ws.cell(row=r + 6, column=cc).border = BORDER
        ws.cell(row=r + 6, column=cc).fill = SUMMARY_TITLE_FILL

    # ---- 行 8：扣重 表头（含 F/G/H 上周期/环比/累计小表头）----
    _set_summary_cell(ws, r + 7, 2, "条件", bold=True, fill=SUMMARY_HEADER_FILL)
    _merge_with_style(ws, r + 7, 2, r + 7, 3)
    ws.cell(row=r + 7, column=3).border = BORDER
    ws.cell(row=r + 7, column=3).fill = SUMMARY_HEADER_FILL
    _set_summary_cell(ws, r + 7, 4, "结果", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 7, 5, "目标值", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 7, 6, "上周期结果", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 7, 7, "环比", bold=True, fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 7, 8, "累计符合率", bold=True, fill=SUMMARY_HEADER_FILL)

    # ---- 行 9：扣重 子表头（含 F9 "扣重符合准确率"）----
    _set_summary_cell(
        ws,
        r + 8,
        2,
        "比值在 0.5~1.5 之间或\n误差绝对值在 150Kg 以内 车次数",
        fill=SUMMARY_HEADER_FILL,
    )
    _merge_with_style(ws, r + 8, 2, r + 8, 3)
    ws.cell(row=r + 8, column=3).border = BORDER
    ws.cell(row=r + 8, column=3).fill = SUMMARY_HEADER_FILL
    _set_summary_cell(ws, r + 8, 4, "扣重符合准确率", fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(
        ws,
        r + 8,
        5,
        "第一阶段 ≥70%\n第二阶段 ≥80%\n第三阶段 ≥90%",
        fill=SUMMARY_HEADER_FILL,
    )
    _merge_with_style(ws, r + 8, 5, r + 9, 5)
    ws.cell(row=r + 9, column=5).border = BORDER
    _set_summary_cell(ws, r + 8, 6, "扣重符合准确率", fill=SUMMARY_HEADER_FILL)
    # G9:G10 留给环比公式（合并）
    # H9:H10 合并 = 累计符合率值
    if p.cumulative_deduction_compliance_rate is not None:
        _set_summary_cell(
            ws, r + 8, 8, float(p.cumulative_deduction_compliance_rate), bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
    else:
        _set_summary_cell(ws, r + 8, 8, "", fill=SUMMARY_HEADER_FILL)
    _merge_with_style(ws, r + 8, 8, r + 9, 8)
    ws.cell(row=r + 9, column=8).border = BORDER
    if p.cumulative_deduction_compliance_rate is None:
        ws.cell(row=r + 9, column=8).fill = SUMMARY_HEADER_FILL

    # ---- 行 10：扣重 数值 + F10 上周期数值 + G9:G10 环比 ----
    _set_summary_cell(ws, r + 9, 2, p.deduction_compliant_count, bold=True)
    _merge_with_style(ws, r + 9, 2, r + 9, 3)
    ws.cell(row=r + 9, column=3).border = BORDER
    formula = f"=IFERROR(B{r + 9}/C{r + 1},0)"
    _set_summary_cell(
        ws, r + 9, 4, formula, bold=True, fill=SUMMARY_RESULT_FILL,
        number_format="0.00%",
    )
    ws.cell(row=r + 9, column=5).border = BORDER
    # F10 上周期扣重符合率
    if p.prev_deduction_compliance_rate is not None:
        _set_summary_cell(
            ws, r + 9, 6, float(p.prev_deduction_compliance_rate), bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
        ratio_formula = f"=IFERROR((D{r + 9}-F{r + 9})/F{r + 9},0)"
        _set_summary_cell(
            ws, r + 8, 7, ratio_formula, bold=True,
            fill=SUMMARY_RESULT_FILL, number_format="0.00%",
        )
        _merge_with_style(ws, r + 8, 7, r + 9, 7)
        ws.cell(row=r + 9, column=7).border = BORDER
    else:
        _set_summary_cell(ws, r + 9, 6, "", fill=SUMMARY_HEADER_FILL)
        _set_summary_cell(ws, r + 8, 7, "/", bold=True, fill=SUMMARY_HEADER_FILL)
        _merge_with_style(ws, r + 8, 7, r + 9, 7)
        ws.cell(row=r + 9, column=7).border = BORDER

    # ---- 行 11：价格差异分布 段落标题 ----
    _set_summary_cell(
        ws, r + 10, 2, "价格差异分布区间（参考值）", bold=True, fill=SUMMARY_TITLE_FILL
    )
    _merge_with_style(ws, r + 10, 2, r + 10, 5)
    for cc in (3, 4, 5):
        ws.cell(row=r + 10, column=cc).border = BORDER
        ws.cell(row=r + 10, column=cc).fill = SUMMARY_TITLE_FILL

    # ---- 行 12：4 档表头 ----
    _set_summary_cell(ws, r + 11, 2, "差价 < 30, 车次数", fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 11, 3, "差价 30~50, 车次数", fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 11, 4, "差价 50~100, 车次数", fill=SUMMARY_HEADER_FILL)
    _set_summary_cell(ws, r + 11, 5, "差价 > 100, 车次数", fill=SUMMARY_HEADER_FILL)

    # ---- 行 13：4 档数值 ----
    _set_summary_cell(ws, r + 12, 2, p.price_diff_lt30, bold=True)
    _set_summary_cell(ws, r + 12, 3, p.price_diff_30_50, bold=True)
    _set_summary_cell(ws, r + 12, 4, p.price_diff_50_100, bold=True)
    _set_summary_cell(ws, r + 12, 5, p.price_diff_gt100, bold=True)

    # ---- 行 14：4 档占比（公式：分母 = 周期内有效检判车次数 C{r+1}）----
    for col_letter, col_idx in zip("BCDE", range(2, 6)):
        formula = f"=IFERROR({col_letter}{r + 12}/C{r + 1},0)"
        _set_summary_cell(
            ws, r + 13, col_idx, formula, fill=SUMMARY_RESULT_FILL,
            number_format="0.00%",
        )

    # ---- F11:H14 合并空白（价格差异分布块没有环比/累计指标）----
    _set_summary_cell(ws, r + 10, 6, "", fill=SUMMARY_HEADER_FILL)
    _merge_with_style(ws, r + 10, 6, r + 13, 8)
    for rr in range(r + 10, r + 14):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).border = BORDER
            ws.cell(row=rr, column=cc).fill = SUMMARY_HEADER_FILL


def _write_cumulative_sheet(
    ws,
    periods: List[PeriodSummary],
    *,
    target_recognition_rate: float,
    target_deduction_compliance_rate: float,
) -> None:
    """写用户截图式「累计统计」页。

    口径：
      · 识别率 = 主料正确且差异≤10%车次 / 周期内有效检判车次
      · 扣重符合率 = 扣重符合车次 / 周期内有效检判车次
    第二个分母刻意与 Sheet1 可见公式保持一致，避免“汇总页”和“概括页”
    对同一指标出现不同结果。
    """
    ws.title = "累计统计"
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 10,
        "B": 28,
        "C": 12,
        "D": 12,
        "E": 13,
        "F": 12,
        "G": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    title = "盛隆检判累计统计"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_CUM_TITLE
        cell.font = Font(bold=True, size=15, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 26

    recognition_title = (
        periods[0].recognition_result_label if periods else "识别准确率"
    )
    headers = [
        (2, 1, "期数"),
        (2, 2, "统计周期"),
        (2, 3, recognition_title),
        (2, 6, "扣重符合率"),
        (3, 3, "总"),
        (3, 4, "对"),
        (3, 5, "率"),
        (3, 6, "对"),
        (3, 7, "率"),
    ]
    for row, col, value in headers:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, size=11, color="FFFFFF" if row == 2 else "000000")
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = FILL_CUM_GROUP if row == 2 else FILL_CUM_HEADER

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    for row in (2, 3):
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = CENTER
            ws.cell(row=row, column=col).border = BORDER

    data_start = 4
    for idx, p in enumerate(periods, start=1):
        row = data_start + idx - 1
        values = [
            f"第{idx}期",
            p.cycle_label,
            p.judgable_trucks,
            p.main_within_10pct_count,
            f"=IFERROR(D{row}/C{row},0)",
            p.deduction_compliant_count,
            f"=IFERROR(F{row}/C{row},0)",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = CENTER
            cell.border = BORDER
            cell.font = Font(size=11)
            if col == 5:
                cell.number_format = "0.00%"
                cell.fill = _rate_fill(
                    p.main_within_10pct_count,
                    p.judgable_trucks,
                    target_recognition_rate,
                )
            elif col == 7:
                cell.number_format = "0.00%"
                cell.fill = _rate_fill(
                    p.deduction_compliant_count,
                    p.judgable_trucks,
                    target_deduction_compliance_rate,
                )
        ws.row_dimensions[row].height = 22

    total_row = data_start + len(periods) + 1
    if periods:
        end_row = data_start + len(periods) - 1
        total_values = [
            "Tol",
            "合计",
            f"=SUM(C{data_start}:C{end_row})",
            f"=SUM(D{data_start}:D{end_row})",
            f"=IFERROR(D{total_row}/C{total_row},0)",
            f"=SUM(F{data_start}:F{end_row})",
            f"=IFERROR(F{total_row}/C{total_row},0)",
        ]
    else:
        total_values = ["Tol", "合计", 0, 0, 0, 0, 0]

    for col, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = FILL_CUM_TOTAL
        cell.font = Font(bold=True, size=11)
        if col in (5, 7):
            cell.number_format = "0.00%"
    ws.row_dimensions[total_row].height = 24


def _write_summary_sheet(ws, periods: List[PeriodSummary]) -> None:
    """把多个周期写成「统计周期概括」sheet（每周期 14 行）"""
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14

    for idx, p in enumerate(periods, start=1):
        base_row = (idx - 1) * PERIOD_BLOCK_ROWS + 1
        ws.row_dimensions[base_row].height = 18  # 行 1 (统计周期)
        ws.row_dimensions[base_row + 4].height = 36  # 行 5 子表头多行
        ws.row_dimensions[base_row + 8].height = 36  # 行 9 扣重条件多行
        _write_one_period_block(ws, base_row, idx, p)


def _write_summary_box(
    ws,
    row: int,
    stats_list: List[DailyShenglongStats],
    target_r: float,
    target_c: float,
) -> int:
    """在 Excel 末尾写汇总框；返回下一行号"""
    total_trucks = sum(s.total_trucks for s in stats_list)
    judgable = sum(s.judgable_trucks for s in stats_list)
    main_same = sum(s.main_same_count for s in stats_list)
    dd_eval = sum(s.deduction_evaluable for s in stats_list)
    dd_ok = sum(s.deduction_compliant_count for s in stats_list)

    r = (main_same / judgable * 100.0) if judgable > 0 else None
    c = (dd_ok / dd_eval * 100.0) if dd_eval > 0 else None

    title = "期间汇总"
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS
    )
    for col in range(1, TOTAL_COLS + 1):
        ws.cell(row=row, column=col).alignment = CENTER
        ws.cell(row=row, column=col).fill = FILL_SUMMARY
        ws.cell(row=row, column=col).border = RED_BORDER

    row += 1
    lines = [
        f"总车数：{total_trucks}    可判定车数：{judgable}    主料一致：{main_same}",
        (
            f"主料识别率 R：{'N/A' if r is None else f'{r:.2f}%'}"
            f"  (目标 ≥ {int(target_r * 100)}%)"
        ),
        f"扣杂可评估车数：{dd_eval}    扣杂符合车数：{dd_ok}",
        (
            f"扣杂符合率：{'N/A' if c is None else f'{c:.2f}%'}"
            f"  (目标 ≥ {int(target_c * 100)}%)"
        ),
    ]
    for line in lines:
        ws.cell(row=row, column=1, value=line).font = Font(bold=True, size=11)
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS
        )
        for col in range(1, TOTAL_COLS + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            ws.cell(row=row, column=col).fill = FILL_SUMMARY
            ws.cell(row=row, column=col).border = RED_BORDER
        row += 1

    return row
