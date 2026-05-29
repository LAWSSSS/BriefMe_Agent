"""Excel writer for accuracy report."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .dict import BINS, MATERIAL_LABEL


def _style_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")


def _apply_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = border


def _write_sheet(ws, title: str, rows):
    ws.title = title
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:H1")
    ws.merge_cells("I1:M1")
    ws.merge_cells("N1:N2")
    ws.merge_cells("O1:S1")

    ws["A1"] = "日期"
    ws["B1"] = "时间"
    ws["C1"] = "烧结类别"
    ws["D1"] = "筛分系统数据"
    ws["I1"] = "视觉系统数据"
    ws["N1"] = "每行误差值"
    ws["O1"] = "各区间误差值"

    for i, b in enumerate(BINS):
        ws.cell(2, 4 + i, f"{b}%")
        ws.cell(2, 9 + i, f"{b}%")
        ws.cell(2, 15 + i, f"{b}%误差值")

    for r in (1, 2):
        for c in range(1, 20):
            _style_header(ws.cell(r, c))

    for i, row in enumerate(rows, start=3):
        t = row["time"]
        if isinstance(t, str):
            from datetime import datetime
            try:
                t = datetime.fromisoformat(t)
            except ValueError:
                t = datetime.strptime(t, "%Y/%m/%d %H:%M")
        ws.cell(i, 1, t.strftime("%Y/%m/%d"))
        ws.cell(i, 2, t.strftime("%H:%M"))
        ws.cell(i, 3, MATERIAL_LABEL)
        for j, b in enumerate(BINS):
            m = round(float(row[f"manual_{b}"]), 2) if row.get(f"manual_{b}") is not None else None
            v = round(float(row[f"visual_{b}"]), 2) if row.get(f"visual_{b}") is not None else None
            e = round(float(row[f"err_{b}"]), 2) if row.get(f"err_{b}") is not None else None
            ws.cell(i, 4 + j, m)
            ws.cell(i, 9 + j, v)
            ws.cell(i, 15 + j, e)
        ws.cell(i, 14, round(float(row["mae"]), 2) if row.get("mae") is not None else None)
        for c in range(1, 20):
            ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="center")

    summary_row = len(rows) + 5
    last_data_row = len(rows) + 2

    ws.cell(summary_row, 1, "五区间平均误差（不包含异常数据）")
    ws.cell(summary_row, 2, f"=ROUND(AVERAGE(D{summary_row},F{summary_row},H{summary_row},J{summary_row},L{summary_row}),2)")
    ws.cell(summary_row, 3, "0-5mm平均误差值（不包含异常数据）")
    ws.cell(summary_row, 4, "=AVERAGE(O:O)")
    ws.cell(summary_row, 5, "5-10mm平均误差值（不包含异常数据）")
    ws.cell(summary_row, 6, "=AVERAGE(P:P)")
    ws.cell(summary_row, 7, "10-25mm平均误差值（不包含异常数据）")
    ws.cell(summary_row, 8, "=AVERAGE(Q:Q)")
    ws.cell(summary_row, 9, "25-40mm平均误差值（不包含异常数据）")
    ws.cell(summary_row, 10, "=AVERAGE(R:R)")
    ws.cell(summary_row, 11, ">40mm平均误差值（不包含异常数据）")
    ws.cell(summary_row, 12, "=AVERAGE(S:S)")

    for c in range(1, 13):
        ws.cell(summary_row, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(summary_row, c).font = Font(bold=False)

    for c in range(1, 20):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["K"].width = 24
    for c in range(4, 20):
        for r in range(3, summary_row + 1):
            ws.cell(r, c).number_format = "0.00"
    _apply_border(ws, 1, summary_row, 1, 19)
    ws.freeze_panes = "A3"


def write_report(output_path: Path, result: Dict[str, Any]) -> None:
    wb = Workbook()
    ws1 = wb.active
    _write_sheet(ws1, "1#烧结矿", result.get("1#", {}).get("rows", []))
    ws2 = wb.create_sheet("2#烧结矿")
    _write_sheet(ws2, "2#烧结矿", result.get("2#", {}).get("rows", []))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
