"""构建检判结果统计 Excel，含图片嵌入。

布局（每辆车 1 个 sheet，每行 5 个板块）：
  Row 1: 表头（无序号）
  Row 2: 数据
  Row 3: 饼图（每两个单元格合并）
  Row 4: 原图 + 判级识别图
  Row 5: 夹杂物图 + 危险物图（第1列为车底图）
后续行：继续 5 个板块一组，直到所有饼图用完。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .page_actions import TABLE_COLUMNS

logger = logging.getLogger(__name__)


def _scale_image(img_path: Path, cell_w_px: int, cell_h_px: int) -> Tuple[int, int]:
    """将图片等比缩放到填满 cell_w_px × cell_h_px（取较严格的一边）。"""
    try:
        with PILImage.open(img_path) as img:
            iw, ih = img.size
    except Exception:
        return cell_w_px, cell_h_px
    scale = min(cell_w_px / iw, cell_h_px / ih)
    return int(iw * scale), int(ih * scale)


def _index_files_by_suffix(directory: Path) -> Dict[int, Path]:
    """索引目录下所有文件，按文件名末尾 _N.ext 提取序号。
    无后缀的文件记为序号 0。
    """
    result: Dict[int, Path] = {}
    if not directory.exists():
        return result
    import re
    for f in sorted(directory.iterdir()):
        if not f.is_file() or f.suffix.lower() not in ('.jpg', '.jpeg', '.png'):
            continue
        m = re.search(r'_(\d+)\.(?:jpg|jpeg|png)$', f.name, re.IGNORECASE)
        if m:
            result[int(m.group(1))] = f
        else:
            # 无后缀 → 序号 0（第一个饼图）
            result.setdefault(0, f)
    return result


def _find_hash(vehicle_dir: Path) -> str:
    """从原图目录中提取 UUID 哈希。"""
    raw_dir = vehicle_dir / "原图"
    if not raw_dir.exists():
        return ""
    import re
    for f in sorted(raw_dir.iterdir()):
        if f.suffix.lower() in ('.jpg', '.jpeg'):
            m = re.search(r'([a-f0-9]{32})', f.name)
            if m:
                return m.group(1)
    return ""


def _format_primary_grade(grade_text: str) -> str:
    """从人工级别文本中提取主料型简称。

    "精炉料-等级二-80%,精炉料-等级一-20%" → "精炉料2"
    "重废-等级一-100%" → "重废1"
    """
    if not grade_text:
        return ""
    import re
    first = grade_text.split(",")[0].strip()
    # 匹配 "料型名-等级X"
    m = re.match(r'^(.+?)-等级([一二三])', first)
    if not m:
        return ""
    material = m.group(1).strip()
    level_map = {"一": "1", "二": "2", "三": "3"}
    level = level_map.get(m.group(2), m.group(2))
    return f"{material}{level}"


def build_stat_excel(
    vehicles: List[Tuple[str, dict, Path]],
    output_path: Path,
) -> Path:
    """为每辆车生成一个工作表，嵌入图片。

    Args:
        vehicles: [(plate, row_data, vehicle_download_dir), ...]
        output_path: 输出 xlsx 路径
    Returns: output_path
    """
    if not vehicles:
        logger.warning("没有车辆数据，跳过 Excel 生成")
        return output_path

    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    section_font = Font(name="微软雅黑", bold=True, size=10, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    img_align = Alignment(horizontal="center", vertical="center")

    for plate, row_data, vehicle_dir in vehicles:
        # 工作表命名：车牌号_人工判级主料型（等级一/二/三 → 1/2/3）
        grade_full = row_data.get("afGradingDef", "")
        grade_short = _format_primary_grade(grade_full)
        sheet_name = f"{plate}_{grade_short}" if grade_short else plate
        # 筛选合法的 sheet 名（openpyxl 要求 <= 31 字符）
        safe_name = sheet_name.replace(":", "").replace("\\", "").replace("/", "").replace("?", "")
        safe_name = safe_name.replace("*", "").replace("[", "").replace("]", "")
        if len(safe_name) > 31:
            safe_name = safe_name[:31]
        ws = wb.create_sheet(title=safe_name)

        # ---- Row 1: 表头 ----
        for col_idx, (_, col_name, _) in enumerate(TABLE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 21

        # ---- Row 2: 数据 ----
        for col_idx, (col_key, _, _) in enumerate(TABLE_COLUMNS, 1):
            value = row_data.get(col_key, "")
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = cell_align
            cell.border = thin_border

        ws.freeze_panes = "A3"

        # ---- 图片区域 ----
        pie_dir = vehicle_dir / "饼图"
        raw_dir = vehicle_dir / "原图"
        grading_dir = vehicle_dir / "判级识别图"
        special_dir = vehicle_dir / "夹杂物图"
        danger_dir = vehicle_dir / "危险物图"

        # 索引饼图
        pie_files: Dict[int, Path] = _index_files_by_suffix(pie_dir)

        # 索引各类型图片
        raw_by_suffix = _index_files_by_suffix(raw_dir)
        grading_by_suffix = _index_files_by_suffix(grading_dir)
        special_by_suffix = _index_files_by_suffix(special_dir)
        danger_by_suffix = _index_files_by_suffix(danger_dir)

        # 车底图（第一个板块的 Row 5 专用）
        file_hash = _find_hash(vehicle_dir)
        finish_img = raw_dir / f"{file_hash}-finish.jpg"

        # 单元格像素常量（1字符≈7px，1磅≈1.333px）
        PIE_W_PX = 42 * 7     # 饼图：2列 × 21字符
        PIE_H_PX = int(110 * 4 / 3)
        IMG_W_PX = 21 * 7     # 其他图：1列 × 21字符
        IMG_H_PX = int(70 * 4 / 3)

        total_pies = len(pie_files)
        blocks_per_row = 5
        cols_per_block = 2  # 每个板块占 2 列

        for block_idx in range(total_pies):
            row_offset = (block_idx // blocks_per_row) * 3  # 每行板块占 3 行
            col_offset = (block_idx % blocks_per_row) * cols_per_block

            base_row = 3 + row_offset
            col_a = 1 + col_offset
            col_b = col_a + 1

            suffix = block_idx + 1

            # ---- Row: 饼图（合并单元格） ----
            ws.merge_cells(start_row=base_row, start_column=col_a, end_row=base_row, end_column=col_b)
            pie_cell = ws.cell(row=base_row, column=col_a)
            pie_key = block_idx

            # 饼图: index=0 无后缀，index>=1 带 _N 后缀
            pie_file = pie_files.get(pie_key)
            if pie_file and pie_file.exists():
                try:
                    w, h = _scale_image(pie_file, PIE_W_PX, PIE_H_PX)
                    img = XLImage(str(pie_file))
                    img.width = w
                    img.height = h
                    pie_cell.alignment = img_align
                    ws.add_image(img, pie_cell.coordinate)
                    ws.row_dimensions[base_row].height = 110
                except Exception as e:
                    logger.warning("插入饼图失败 %s: %s", pie_file.name, e)

            # ---- Row + 1: 原图 + 判级识别图 ----
            img_row = base_row + 1

            raw_file = raw_by_suffix.get(suffix)
            if raw_file and raw_file.exists():
                try:
                    w, h = _scale_image(raw_file, IMG_W_PX, IMG_H_PX)
                    img = XLImage(str(raw_file))
                    img.width = w
                    img.height = h
                    cell = ws.cell(row=img_row, column=col_a)
                    cell.alignment = img_align
                    ws.add_image(img, cell.coordinate)
                except Exception as e:
                    logger.warning("插入原图失败 %s: %s", raw_file.name, e)

            grading_file = grading_by_suffix.get(suffix)
            if grading_file and grading_file.exists():
                try:
                    w, h = _scale_image(grading_file, IMG_W_PX, IMG_H_PX)
                    img = XLImage(str(grading_file))
                    img.width = w
                    img.height = h
                    cell = ws.cell(row=img_row, column=col_b)
                    cell.alignment = img_align
                    ws.add_image(img, cell.coordinate)
                except Exception as e:
                    logger.warning("插入判级识别图失败 %s: %s", grading_file.name, e)

            ws.row_dimensions[img_row].height = 70

            # ---- Row + 2: 夹杂物图 + 危险物图（第一列为车底图） ----
            special_row = base_row + 2

            if block_idx == 0 and finish_img.exists():
                # 第一个板块：车底图
                try:
                    w, h = _scale_image(finish_img, IMG_W_PX, IMG_H_PX)
                    img = XLImage(str(finish_img))
                    img.width = w
                    img.height = h
                    cell = ws.cell(row=special_row, column=col_b)
                    cell.alignment = img_align
                    ws.add_image(img, cell.coordinate)
                except Exception as e:
                    logger.warning("插入车底图失败: %s", e)
            else:
                special_file = special_by_suffix.get(suffix)
                if special_file and special_file.exists():
                    try:
                        w, h = _scale_image(special_file, IMG_W_PX, IMG_H_PX)
                        img = XLImage(str(special_file))
                        img.width = w
                        img.height = h
                        cell = ws.cell(row=special_row, column=col_a)
                        cell.alignment = img_align
                        ws.add_image(img, cell.coordinate)
                    except Exception as e:
                        logger.warning("插入夹杂物图失败 %s: %s", special_file.name, e)

            danger_file = danger_by_suffix.get(suffix)
            if danger_file and danger_file.exists():
                try:
                    w, h = _scale_image(danger_file, IMG_W_PX, IMG_H_PX)
                    img = XLImage(str(danger_file))
                    img.width = w
                    img.height = h
                    cell = ws.cell(row=special_row, column=col_b)
                    cell.alignment = img_align
                    ws.add_image(img, cell.coordinate)
                except Exception as e:
                    logger.warning("插入危险物图失败 %s: %s", danger_file.name, e)

            ws.row_dimensions[special_row].height = 70

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("统计 Excel 已保存: %s (%d 辆车)", output_path, len(vehicles))
    return output_path
