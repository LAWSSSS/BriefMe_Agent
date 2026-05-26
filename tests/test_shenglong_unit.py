"""盛隆废钢模块单元/组件 smoke 测试（不需要访问真实 API）

运行：
  /opt/anaconda3/bin/python tests/test_shenglong_unit.py
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from agent.shenglong.calculator import (
    aggregate_daily,
    aggregate_period,
    aggregate_period_heavy_normalized,
    calc_truck,
)
from agent.shenglong.dict import (
    EMPTY_TYPES,
    REFERENCE_TYPES,
    STEEL_TYPE,
    filter_main_candidates,
    get_main_type_from_list,
    get_material_name,
    is_valid,
)
from agent.shenglong.excel_writer import write_stats_xlsx
from agent.shenglong.models import (
    ManualOperator,
    MaterialRate,
    ShenglongRecord,
    TruckStat,
)
from config.settings import settings


def test_dict():
    # 覆盖全部 17 个条目（0 + 1..16）
    assert len(STEEL_TYPE) == 17
    assert STEEL_TYPE[0] == ""
    assert STEEL_TYPE[11] == "中废"
    assert STEEL_TYPE[14] == "钢筋切粒"
    assert STEEL_TYPE[16] == "超标"

    assert get_material_name(1) == "重废1"
    assert get_material_name(0) == "--"
    assert get_material_name(None) == "--"
    assert get_material_name(16) == "超标"

    # 参考型/空值判定
    assert 16 in REFERENCE_TYPES
    assert 0 in EMPTY_TYPES
    assert not is_valid(0)
    assert not is_valid(16)
    assert is_valid(1)
    assert is_valid(11)

    # 剔除超标/空值之后选主料
    items = [(0, 0.1), (11, 0.3), (16, 0.5), (12, 0.2)]  # 16 超标占 50%
    cands = filter_main_candidates(items)
    assert (16, 0.5) not in cands
    assert (0, 0.1) not in cands
    main = get_main_type_from_list(items)
    assert main == (11, 0.3)  # 中废赢

    # 只有超标时返回 None（没合法主料）
    assert get_main_type_from_list([(16, 0.9), (0, 0.1)]) is None
    print("dict OK")


def test_shenglong_record_station_number_list():
    """列表接口 stationNumber 可能返回 list，不能直接 int(list) 崩溃。"""
    base = {
        "flowCode": "f1",
        "carNumber": "桂TEST",
        "createTime": "2026-04-14 10:00:00",
    }
    r1 = ShenglongRecord.from_list_item({**base, "stationNumber": [36]})
    assert r1.station_number == 36

    r2 = ShenglongRecord.from_list_item({**base, "stationNumber": [36, 53]})
    assert r2.station_number == "36/53"

    r3 = ShenglongRecord.from_list_item({**base, "stationNumber": []})
    assert r3.station_number == 0

    r4 = ShenglongRecord.from_list_item({**base, "stationNumber": "36"})
    assert r4.station_number == 36
    print("shenglong record station list OK")


def _sample_detail_same_main():
    """双方主料都是 重废1（code=1）→ 一致；扣重在 0.5~1.5 区间"""
    return {
        "manualCheckResultVO": {
            "avgResult": [
                {"steelType": 1, "steelRate": 0.80},
                {"steelType": 11, "steelRate": 0.20},
            ],
            "avgDeduction": 0.120,  # 吨
            "avgSteelPrice": 2800,
            "checkDetails": [
                {
                    "operatorName": "张三",
                    "deduction": 0.115,
                    "steelPrice": 2810,
                    "details": [
                        {"steelType": 1, "steelRate": 0.80},
                        {"steelType": 11, "steelRate": 0.20},
                    ],
                },
                {
                    "operatorName": "李四",
                    "deduction": 0.120,
                    "steelPrice": 2790,
                    "details": [
                        {"steelType": 1, "steelRate": 0.82},
                        {"steelType": 11, "steelRate": 0.18},
                    ],
                },
                {
                    "operatorName": "王五",
                    "deduction": 0.125,
                    "steelPrice": 2800,
                    "details": [
                        {"steelType": 1, "steelRate": 0.78},
                        {"steelType": 11, "steelRate": 0.22},
                    ],
                },
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 1, "steelRate": 0.75},  # 80% vs 75% → diff 5%
                {"steelType": 11, "steelRate": 0.25},
            ],
            "totalDeductWeight": 140,  # kg → 0.14 吨 vs 人工 0.12 吨 → ratio 1.167, diff 0.02
        },
    }


def _sample_detail_mismatch():
    """人工主料=重废1；AI 主料=中废 → 不一致

    人工 1 个有效检判员（不在黑名单里），用于覆盖"剔除黑名单后仍有人"的场景。
    """
    return {
        "manualCheckResultVO": {
            "avgDeduction": 0.20,
            "avgSteelPrice": 2800,
            "checkDetails": [
                {
                    "operatorName": "张三",
                    "deduction": 0.20,
                    "steelPrice": 2800,
                    "details": [
                        {"steelType": 1, "steelRate": 0.60},
                        {"steelType": 11, "steelRate": 0.40},
                    ],
                },
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 11, "steelRate": 0.70},
                {"steelType": 1, "steelRate": 0.30},
            ],
            "totalDeductWeight": 500,  # 0.5t, 人工 0.2t, ratio 2.5, diff 0.3 → 均不符合 → 扣杂 False
        },
    }


def _sample_detail_chaobiao_heavy():
    """超标占比最高，但按规则应当被剔除 → 主料应该是 重废1（1 个非黑名单 operator）"""
    return {
        "manualCheckResultVO": {
            "avgDeduction": 0.11,
            "avgSteelPrice": 2800,
            "checkDetails": [
                {
                    "operatorName": "李四",
                    "deduction": 0.11,
                    "steelPrice": 2800,
                    "details": [
                        {"steelType": 16, "steelRate": 0.55},
                        {"steelType": 1, "steelRate": 0.30},
                        {"steelType": 11, "steelRate": 0.15},
                    ],
                },
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 16, "steelRate": 0.40},
                {"steelType": 1, "steelRate": 0.45},
                {"steelType": 11, "steelRate": 0.15},
            ],
            "totalDeductWeight": 115,
        },
    }


def test_calc():
    cfg = settings.shenglong

    # 一致
    t1 = calc_truck("2026-04-22", "鲁A-0001", 1, "flow-1",
                    _sample_detail_same_main(), cfg)
    assert t1.manual_main.steel_type == 1
    assert t1.ai_main.steel_type == 1
    assert t1.main_same is True
    # 差异值 = |80 - 75| = 5.0
    assert abs(t1.diff_rate - 5.0) < 1e-6
    assert t1.manual_deduct_ton == 0.120
    assert abs(t1.ai_deduct_ton - 0.140) < 1e-9
    # ratio 0.14/0.12 ≈ 1.167，在 [0.5, 1.5] 区间 → 扣杂合格
    assert t1.deduction_compliant is True
    assert len(t1.manual_operators) == 3
    assert t1.manual_operators[0].name == "张三"
    assert t1.manual_operators[0].main.steel_type == 1

    # 不一致
    t2 = calc_truck("2026-04-22", "鲁A-0002", 2, "flow-2",
                    _sample_detail_mismatch(), cfg)
    assert t2.manual_main.steel_type == 1
    assert t2.ai_main.steel_type == 11
    assert t2.main_same is False
    assert t2.diff_rate is None  # 主料不一致时差异值 /
    # ratio 2.5 不在 [0.5,1.5]，|diff|=0.3 > 0.15 → 不合格
    assert t2.deduction_compliant is False
    # 新规则：人工合并基于 checkDetails 后的 operator 列表（已剔除黑名单 5 人）
    # 这里 1 个非黑名单 operator → manual_operators 长度 = 1
    assert len(t2.manual_operators) == 1
    assert t2.manual_operators[0].name == "张三"

    # 超标剔除
    t3 = calc_truck("2026-04-22", "鲁A-0003", 3, "flow-3",
                    _sample_detail_chaobiao_heavy(), cfg)
    assert t3.manual_main.steel_type == 1  # 16 被剔除，剩下 1(30%) 和 11(15%) → 1 胜
    assert t3.ai_main.steel_type == 1  # 16 被剔除，剩下 1(45%) 和 11(15%) → 1 胜
    # 差异值 = |30 - 45| = 15，仍然计算出来
    assert abs(t3.diff_rate - 15.0) < 1e-6
    # 新规则：料型一致但差异 15% > 10% → 主料型不正确
    assert t3.main_same is False

    print("calc OK")
    print(f"  t1.diff_rate={t1.diff_rate}, t1.weight_ratio={t1.weight_ratio:.3f}")
    print(f"  t2.deduction_compliant={t2.deduction_compliant}")
    print(f"  t3.main_same={t3.main_same}")


def _build_sample_trucks():
    """供 test_aggregate_and_summary / test_excel 复用的真实 TruckStat 列表"""
    cfg = settings.shenglong
    t1 = calc_truck("2026-04-22", "鲁A-0001", 1, "flow-1",
                    _sample_detail_same_main(), cfg)
    t2 = calc_truck("2026-04-22", "鲁A-0002", 2, "flow-2",
                    _sample_detail_mismatch(), cfg)
    t3 = calc_truck("2026-04-22", "鲁A-0003", 3, "flow-3",
                    _sample_detail_chaobiao_heavy(), cfg)
    return [t1, t2, t3]


def test_aggregate_and_summary():
    cfg = settings.shenglong
    trucks = _build_sample_trucks()
    stats = aggregate_daily("2026-04-22", trucks)

    assert stats.total_trucks == 3
    assert stats.judgable_trucks == 3  # 三辆车人工 AI 主料都非空
    # 新规则：t1 diff=5% ≤ 10% → 正确；t3 diff=15% > 10% → 不正确
    assert stats.main_same_count == 1  # 仅 t1
    # 识别率 = 1/3 = 33.33%
    assert abs(stats.recognition_rate - 1 / 3 * 100.0) < 1e-6
    # 扣杂评估：3 辆车都有扣重 → 3；合格：t1、t3（t3 diff 0.005 < 0.15 → ok）
    assert stats.deduction_evaluable == 3
    assert stats.deduction_compliant_count == 2
    assert abs(stats.deduction_compliance_rate - 2 / 3 * 100.0) < 1e-6

    text = stats.summary_text(
        target_recognition_rate=cfg.target_recognition_rate,
        target_deduction_compliance_rate=cfg.target_deduction_compliance_rate,
    )
    print("summary_text:\n" + text)
    assert "2026年4月22日" in text
    assert "3 车" in text
    assert "正确 1/3 辆" in text
    assert "未达标" in text  # 33.33 < 92 且 66.67 < 90
    print("aggregate+summary OK")


def test_excel():
    cfg = settings.shenglong
    trucks = _build_sample_trucks()
    day1 = aggregate_daily("2026-04-22", [trucks[0], trucks[1]])
    day2 = aggregate_daily("2026-04-23", [trucks[2]])

    out = Path("downloads/shenglong/_unit_test/report.xlsx")
    write_stats_xlsx(
        [day1, day2],
        out,
        target_recognition_rate=cfg.target_recognition_rate,
        target_deduction_compliance_rate=cfg.target_deduction_compliance_rate,
    )
    assert out.exists() and out.stat().st_size > 2000
    print(f"excel OK → {out} ({out.stat().st_size} bytes)")


def test_period_summary_aggregation():
    """跨日聚合到 PeriodSummary：覆盖 4 档价格分桶 + 主料正确两种判定口径"""
    trucks = _build_sample_trucks()
    # 给三辆车补一个价格差异（calc_truck 已经算了 price_diff = |AI-人工|）
    # t1: AI 单价 None? → 看 calc_truck 实现, AI steelPrice 字段没传 → t1.price_diff = None
    # 这里手动注入便于测分桶
    trucks[0].price_diff = 10.0   # < 30
    trucks[1].price_diff = 60.0   # 50~100
    trucks[2].price_diff = 120.0  # > 100

    day1 = aggregate_daily("2026-04-22", trucks[:2])
    day2 = aggregate_daily("2026-04-23", trucks[2:])

    period = aggregate_period([day1, day2], "2026-04-22", "2026-04-23")
    assert period.cycle_label == "2026.4.22 至 2026.4.23"
    assert period.judgable_trucks == 3
    # 名字一致：t1（同 1）、t3（同 1）→ 2；t2 名字不同 → False
    assert period.main_name_match_count == 2
    # 名字一致 AND 差异≤10%：仅 t1（5%）→ 1；t3 是 15% 不算
    assert period.main_within_10pct_count == 1
    assert abs(period.recognition_rate_pct - 1 / 3 * 100.0) < 1e-6
    # 扣杂：t1 / t3 合格，t2 不合格 → 2/3
    assert period.deduction_compliant_count == 2
    assert period.deduction_evaluable == 3
    assert abs(period.deduction_compliance_rate_pct - 2 / 3 * 100.0) < 1e-6
    # 价格分桶
    assert period.price_diff_lt30 == 1
    assert period.price_diff_30_50 == 0
    assert period.price_diff_50_100 == 1
    assert period.price_diff_gt100 == 1
    assert period.price_diff_evaluable == 3
    print(f"period OK → recognition={period.recognition_rate_pct:.2f}%")


def test_excel_with_summary_sheet():
    """xlsx 应该有三个 sheet：统计周期概括 + 累计统计 + 检判统计详情"""
    from openpyxl import load_workbook

    trucks = _build_sample_trucks()
    for t, pd in zip(trucks, [10.0, 60.0, 120.0]):
        t.price_diff = pd

    day1 = aggregate_daily("2026-04-22", trucks[:2])
    day2 = aggregate_daily("2026-04-23", trucks[2:])
    period = aggregate_period([day1, day2], "2026-04-22", "2026-04-23")

    out = Path("downloads/shenglong/_unit_test/report_with_summary.xlsx")
    write_stats_xlsx([day1, day2], out, period_summary=period)
    assert out.exists()

    wb = load_workbook(str(out))
    assert wb.sheetnames == ["统计周期概括", "累计统计", "检判统计详情"]

    ws = wb["统计周期概括"]
    # row 1 周期标签
    assert ws.cell(row=1, column=2).value == "统计周期"
    assert "2026.4.22" in str(ws.cell(row=1, column=3).value)
    # row 2 周期内有效检判车次数
    assert ws.cell(row=2, column=3).value == 3
    # row 6 主料型相同/差异<10% 数值
    assert ws.cell(row=6, column=2).value == 2
    assert ws.cell(row=6, column=3).value == 1
    # row 6 D 列写入的是公式
    assert str(ws.cell(row=6, column=4).value).startswith("=IFERROR(")
    # row 13 价格分桶
    assert ws.cell(row=13, column=2).value == 1
    assert ws.cell(row=13, column=3).value == 0
    assert ws.cell(row=13, column=4).value == 1
    assert ws.cell(row=13, column=5).value == 1

    ws_cum = wb["累计统计"]
    assert ws_cum.cell(row=1, column=1).value == "盛隆检判累计统计"
    assert ws_cum.cell(row=2, column=3).value == "识别准确率"
    assert ws_cum.cell(row=4, column=1).value == "第1期"
    assert ws_cum.cell(row=4, column=3).value == period.judgable_trucks
    assert ws_cum.cell(row=4, column=4).value == period.main_within_10pct_count
    assert ws_cum.cell(row=4, column=5).value == "=IFERROR(D4/C4,0)"
    assert ws_cum.cell(row=4, column=6).value == period.deduction_compliant_count
    assert ws_cum.cell(row=4, column=7).value == "=IFERROR(F4/C4,0)"
    assert ws_cum.cell(row=6, column=1).value == "Tol"
    assert ws_cum.cell(row=6, column=5).value == "=IFERROR(D6/C6,0)"
    print("summary sheet OK")


def test_operator_from_dict():
    d = {
        "operatorName": "赵六",
        "deduction": 0.115,
        "steelPrice": 2800,
        "details": [
            {"steelType": 16, "steelRate": 0.5},  # 超标：应剔除
            {"steelType": 11, "steelRate": 0.3},
            {"steelType": 0, "steelRate": 0.2},  # 空值：应剔除
        ],
    }
    op = ManualOperator.from_dict(d)
    assert op.name == "赵六"
    assert op.deduction_ton == 0.115
    assert op.main is not None
    assert op.main.steel_type == 11  # 剔除 16 和 0 后剩下 11
    print("operator_from_dict OK")


def test_normalize_deduction_kg_to_ton():
    """单位自动修正：> 10 视为 kg → /1000 转吨"""
    from agent.shenglong.models import _normalize_deduction

    # 正常吨值不动
    assert _normalize_deduction(0.0) == 0.0
    assert _normalize_deduction(0.115) == 0.115
    assert _normalize_deduction(2.28) == 2.28
    assert _normalize_deduction(5.0) == 5.0          # 仍按吨认（虽然偏大）
    assert _normalize_deduction(10.0) == 10.0        # 阈值不含等号

    # 异常值 > 10 → 当 kg 处理
    assert _normalize_deduction(26.0) == 0.026       # 截图里王高翔的 26.000
    assert _normalize_deduction(35.0) == 0.035       # 截图里王高翔的 35.000
    assert _normalize_deduction(150.0) == 0.15
    assert _normalize_deduction(2280.0) == 2.28
    assert _normalize_deduction(2280.0, "陈宇辉") == 2.28

    # None / 非数值
    assert _normalize_deduction(None) is None
    assert _normalize_deduction("abc") is None
    print("normalize_deduction OK")


def test_operator_from_dict_kg_correction():
    """from_dict 路径：录入 2280（kg）应被自动修正为 2.28（吨）"""
    d = {
        "operatorName": "陈宇辉",
        "deduction": 2280.000,  # ← 用户截图里的真实异常值
        "steelPrice": 2800,
        "details": [
            {"steelType": 1, "steelRate": 0.80},
            {"steelType": 11, "steelRate": 0.20},
        ],
    }
    op = ManualOperator.from_dict(d)
    assert op.name == "陈宇辉"
    assert op.deduction_ton == 2.28, (
        f"期望自动修正为 2.28（吨），实际 {op.deduction_ton}"
    )
    print("operator_kg_correction OK")


def test_calc_truck_with_kg_input():
    """端到端：3 个 operator 中一人录入 kg，平均后扣重应符合常识"""
    from agent.shenglong.calculator import calc_truck

    detail = {
        "manualCheckResultVO": {
            "avgResult": [
                {"steelType": 1, "avgRate": 0.80},
                {"steelType": 11, "avgRate": 0.20},
            ],
            "avgDeduction": 0.31,
            "avgSteelPrice": 2800,
            "checkDetails": [
                {
                    "operatorName": "王冰磊",
                    "deduction": 0.350,  # 正常吨
                    "steelPrice": 2800,
                    "details": [{"steelType": 1, "steelRate": 0.80}],
                },
                {
                    "operatorName": "陈宇辉",
                    "deduction": 2280.000,  # ← kg 录入异常，应被修正为 2.28
                    "steelPrice": 2800,
                    "details": [{"steelType": 1, "steelRate": 0.80}],
                },
                {
                    "operatorName": "王高翔",
                    "deduction": 0.300,
                    "steelPrice": 2800,
                    "details": [{"steelType": 1, "steelRate": 0.80}],
                },
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [{"steelType": 1, "avgRate": 0.85}],
            "totalDeductWeight": 350,  # kg → 0.35 吨
        },
    }
    truck = calc_truck(
        date_str="2026-04-22",
        car_number="渝AKG001",
        station_number=1,
        flow_code="kg-test",
        detail_data=detail,
        cfg=settings.shenglong,
    )

    # 关键断言 1：单个 operator 的扣重（Excel 人工详情②列那一格）已被修正
    op_chen = next(op for op in truck.manual_operators if op.name == "陈宇辉")
    assert op_chen.deduction_ton == 2.28, (
        f"陈宇辉扣重应自动修正为 2.28，实际 {op_chen.deduction_ton}"
    )

    # 关键断言 2：人工平均扣重 = (0.350 + 2.28 + 0.300) / 3 = 0.9767（吨）
    # 而不是修正前的 (0.350 + 2280 + 0.300) / 3 = 760.22（明显荒谬）
    expected_avg = (0.350 + 2.28 + 0.300) / 3
    assert abs(truck.manual_deduct_ton - expected_avg) < 1e-6, (
        f"人工平均扣重应 ≈ {expected_avg:.4f}，实际 {truck.manual_deduct_ton}"
    )
    print("calc_truck_with_kg_input OK")


# =====================================================================
# 新规则：剔除指定检判员 + 任一方缺失不计入统计
# =====================================================================
def _sample_with_excluded_operators():
    """3 个检判员里有 1 个在黑名单（施宏波），应被剔除后用剩 2 人合并"""
    return {
        "manualCheckResultVO": {
            "avgDeduction": 0.120,
            "avgSteelPrice": 2800,
            "checkDetails": [
                {
                    "operatorName": "张三",
                    "deduction": 0.115,
                    "steelPrice": 2810,
                    "details": [
                        {"steelType": 1, "steelRate": 0.80},
                        {"steelType": 11, "steelRate": 0.20},
                    ],
                },
                {
                    "operatorName": "施宏波",  # 黑名单 → 应被剔除
                    "deduction": 999.0,  # 故意放离群值，验证被剔除
                    "steelPrice": 9999,
                    "details": [
                        {"steelType": 11, "steelRate": 0.99},
                        {"steelType": 1, "steelRate": 0.01},
                    ],
                },
                {
                    "operatorName": "王五",
                    "deduction": 0.125,
                    "steelPrice": 2800,
                    "details": [
                        {"steelType": 1, "steelRate": 0.78},
                        {"steelType": 11, "steelRate": 0.22},
                    ],
                },
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 1, "steelRate": 0.75},
                {"steelType": 11, "steelRate": 0.25},
            ],
            "totalDeductWeight": 140,
        },
    }


def _sample_all_operators_excluded():
    """全部检判员都在黑名单 → 整车人工缺失"""
    return {
        "manualCheckResultVO": {
            "avgDeduction": 0.20,
            "avgSteelPrice": 2800,
            "checkDetails": [
                {"operatorName": "施宏波", "deduction": 0.20, "steelPrice": 2800,
                 "details": [{"steelType": 1, "steelRate": 0.60}]},
                {"operatorName": "冉星明", "deduction": 0.21, "steelPrice": 2800,
                 "details": [{"steelType": 1, "steelRate": 0.55}]},
                {"operatorName": "周倩", "deduction": 0.22, "steelPrice": 2800,
                 "details": [{"steelType": 1, "steelRate": 0.65}]},
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 1, "steelRate": 0.70},
            ],
            "totalDeductWeight": 200,
        },
    }


def _sample_only_ai_no_manual():
    """checkDetails 完全为空 → 人工缺失，但 AI 有结果"""
    return {
        "manualCheckResultVO": {"checkDetails": []},
        "totalCheckResult": {
            "steelTypeRateList": [
                {"steelType": 1, "steelRate": 0.80},
            ],
            "totalDeductWeight": 150,
        },
    }


def _sample_only_manual_no_ai():
    """AI 缺主料（steelTypeRateList 为空），人工有 1 个 operator"""
    return {
        "manualCheckResultVO": {
            "avgDeduction": 0.20,
            "checkDetails": [
                {"operatorName": "张三", "deduction": 0.20, "steelPrice": 2800,
                 "details": [{"steelType": 1, "steelRate": 0.80}]},
            ],
        },
        "totalCheckResult": {
            "steelTypeRateList": [],
            "totalDeductWeight": None,
        },
    }


def test_excluded_operators_filtered():
    """3 人中有 1 人是黑名单 → 应被剔除，剩 2 人重新平均"""
    cfg = settings.shenglong
    t = calc_truck(
        "2026-04-22", "鲁A-0010", 1, "flow-x",
        _sample_with_excluded_operators(), cfg,
    )

    # 1) 黑名单不进 manual_operators
    assert len(t.manual_operators) == 2
    names = [op.name for op in t.manual_operators]
    assert "施宏波" not in names
    assert set(names) == {"张三", "王五"}

    # 2) 主料合并基于剩下的人，离群值 99% 中废没参与平均
    # 张三/王五: 重废1 = (80 + 78)/2 = 79；中废 = (20 + 22)/2 = 21
    assert t.manual_main is not None
    assert t.manual_main.steel_type == 1
    assert abs(t.manual_main.rate - 79.0) < 1e-6

    # 3) 扣重平均 = (0.115 + 0.125) / 2 = 0.120（黑名单 999 没参与）
    assert abs(t.manual_deduct_ton - 0.120) < 1e-9

    # 4) 单价平均 = (2810 + 2800) / 2 = 2805
    assert abs(t.manual_steel_price - 2805.0) < 1e-6

    print("excluded_operators OK")


def test_all_operators_excluded_means_manual_missing():
    """全员都在黑名单 → 整车人工缺失，不计入识别率/扣重符合率统计"""
    cfg = settings.shenglong
    t = calc_truck(
        "2026-04-22", "鲁A-0011", 1, "flow-y",
        _sample_all_operators_excluded(), cfg,
    )

    assert len(t.manual_operators) == 0
    assert t.manual_main is None
    assert t.manual_materials == []
    assert t.manual_deduct_ton is None
    assert t.manual_steel_price is None

    # 关键：缺失 → main_same=None / deduction_compliant=None
    assert t.main_same is None
    assert t.main_name_match is None
    assert t.deduction_compliant is None

    # 聚合时不计入分母
    stats = aggregate_daily("2026-04-22", [t])
    assert stats.judgable_trucks == 0
    assert stats.deduction_evaluable == 0
    assert stats.recognition_rate is None
    assert stats.deduction_compliance_rate is None
    print("all_excluded OK")


def test_one_side_missing_not_counted():
    """人工或 AI 任一缺失 → 不计入统计"""
    cfg = settings.shenglong

    t_no_manual = calc_truck(
        "2026-04-22", "鲁A-0012", 1, "flow-z1",
        _sample_only_ai_no_manual(), cfg,
    )
    assert t_no_manual.manual_main is None
    assert t_no_manual.ai_main is not None
    assert t_no_manual.main_same is None
    assert t_no_manual.deduction_compliant is None

    t_no_ai = calc_truck(
        "2026-04-22", "鲁A-0013", 1, "flow-z2",
        _sample_only_manual_no_ai(), cfg,
    )
    assert t_no_ai.manual_main is not None
    assert t_no_ai.ai_main is None
    assert t_no_ai.main_same is None
    assert t_no_ai.deduction_compliant is None

    # 跨日聚合：两辆车都有展示但不进分母
    stats = aggregate_daily("2026-04-22", [t_no_manual, t_no_ai])
    assert stats.total_trucks == 2
    assert stats.judgable_trucks == 0
    assert stats.deduction_evaluable == 0
    print("one_side_missing OK")


def test_excel_summary_sheet1_with_prev_period():
    """Sheet1 F/G 列：当 prev_recognition_rate 提供时写入数值 + 环比公式"""
    from openpyxl import load_workbook
    from agent.shenglong.models import PeriodSummary

    trucks = _build_sample_trucks()
    for t, pd in zip(trucks, [10.0, 60.0, 120.0]):
        t.price_diff = pd

    day1 = aggregate_daily("2026-04-22", trucks[:2])
    day2 = aggregate_daily("2026-04-23", trucks[2:])
    period = aggregate_period([day1, day2], "2026-04-22", "2026-04-23")

    # 模拟上周期：识别率 6.82%，扣重符合率 20.45%（参考表里 reference 的真实数）
    period.prev_recognition_rate = 0.0681818
    period.prev_deduction_compliance_rate = 0.2045454
    period.prev_cycle_label = "2026-04-14 ~ 2026-04-22"

    out = Path("downloads/shenglong/_unit_test/report_with_prev.xlsx")
    write_stats_xlsx([day1, day2], out, period_summary=period)
    assert out.exists()

    wb = load_workbook(str(out))
    ws = wb["统计周期概括"]
    # F1 = "指标变化"
    assert ws.cell(row=1, column=6).value == "指标变化"
    # F4/G4 子表头
    assert ws.cell(row=4, column=6).value == "上周期结果"
    assert ws.cell(row=4, column=7).value == "环比"
    # F5 = "识别准确率"
    assert ws.cell(row=5, column=6).value == "识别准确率"
    # F6 = 上周期识别率（小数）
    assert abs(float(ws.cell(row=6, column=6).value) - 0.0681818) < 1e-6
    # G5 = 环比公式
    assert str(ws.cell(row=5, column=7).value).startswith("=IFERROR(")
    assert "F6" in str(ws.cell(row=5, column=7).value)
    # F10 = 上周期扣重符合率
    assert abs(float(ws.cell(row=10, column=6).value) - 0.2045454) < 1e-6
    # G9 = 扣重环比公式
    assert str(ws.cell(row=9, column=7).value).startswith("=IFERROR(")
    print("sheet1 with prev OK")


def test_master_xlsx_two_cycles_with_auto_prev_chain():
    """多周期主表：2 个周期，Sheet1 应有 28 行，Sheet2 应有 2 段。
    第 2 周期的 prev 自动从第 1 周期取值。"""
    from openpyxl import load_workbook
    from agent.shenglong.excel_writer import write_master_xlsx

    # 周期 1：4.14-4.22
    trucks_a = _build_sample_trucks()
    for t, pd in zip(trucks_a, [10.0, 60.0, 120.0]):
        t.price_diff = pd
    day1a = aggregate_daily("2026-04-14", trucks_a[:2])
    day2a = aggregate_daily("2026-04-15", trucks_a[2:])
    period_a = aggregate_period([day1a, day2a], "2026-04-14", "2026-04-22")

    # 周期 2：4.23-4.29
    trucks_b = _build_sample_trucks()
    for t, pd in zip(trucks_b, [25.0, 80.0, 5.0]):
        t.price_diff = pd
    day1b = aggregate_daily("2026-04-23", trucks_b[:1])
    day2b = aggregate_daily("2026-04-24", trucks_b[1:])
    period_b = aggregate_period([day1b, day2b], "2026-04-23", "2026-04-29")

    cycles = [
        ([day1a, day2a], period_a),
        ([day1b, day2b], period_b),
    ]

    out = Path("downloads/shenglong/_unit_test/master_two_cycles.xlsx")
    write_master_xlsx(cycles, out)
    assert out.exists()

    wb = load_workbook(str(out))
    assert wb.sheetnames == ["统计周期概括", "累计统计", "检判统计详情"]

    # ---- Sheet1：两个 14 行块依次往下 ----
    ws_s = wb["统计周期概括"]
    # 块 1（行 1~14）：A1=1, C1=周期 1 标签
    assert ws_s.cell(row=1, column=1).value == 1
    assert "2026.4.14" in str(ws_s.cell(row=1, column=3).value)
    # 块 2（行 15~28）：A15=2, C15=周期 2 标签
    assert ws_s.cell(row=15, column=1).value == 2
    assert "2026.4.23" in str(ws_s.cell(row=15, column=3).value)
    # 块 1 首期：F6 留空 / G5 = "/"
    assert ws_s.cell(row=6, column=6).value in (None, "")
    assert ws_s.cell(row=5, column=7).value == "/"
    # 块 2：F20（=15+5）应该被自动填上 = 周期 1 的识别率
    f20 = ws_s.cell(row=20, column=6).value
    assert f20 is not None and f20 != ""
    expected_prev = period_a.recognition_rate_pct / 100.0
    assert abs(float(f20) - expected_prev) < 1e-9
    # 块 2：G19 应是环比公式（G19:G20 合并）
    g19 = str(ws_s.cell(row=19, column=7).value)
    assert g19.startswith("=IFERROR(")

    # ---- Sheet2：两段，每段含周期标题 ----
    ws_d = wb["检判统计详情"]
    # 第 1 行整张表标题
    assert "盛隆赛迪废钢判级" in str(ws_d.cell(row=1, column=1).value)
    # 收集 A 列所有非空文字，应能找到 2 个段标题（"第 N 期 ..."）
    period_titles = []
    for r in range(1, ws_d.max_row + 1):
        v = ws_d.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith("第 "):
            period_titles.append(v)
    assert len(period_titles) == 2
    assert "第 1 期" in period_titles[0]
    assert "第 2 期" in period_titles[1]
    print("master xlsx two cycles OK")


def test_master_xlsx_user_provided_prev_not_overridden():
    """如果用户已经手动设置了 prev_recognition_rate，自动链不应覆盖"""
    from agent.shenglong.excel_writer import write_master_xlsx

    trucks_a = _build_sample_trucks()
    day_a = aggregate_daily("2026-04-14", trucks_a)
    period_a = aggregate_period([day_a], "2026-04-14", "2026-04-22")

    trucks_b = _build_sample_trucks()
    day_b = aggregate_daily("2026-04-23", trucks_b)
    period_b = aggregate_period([day_b], "2026-04-23", "2026-04-29")
    # 用户手动设了一个 prev（比如基于历史人工记录）
    period_b.prev_recognition_rate = 0.5  # 50%

    cycles = [
        ([day_a], period_a),
        ([day_b], period_b),
    ]
    out = Path("downloads/shenglong/_unit_test/master_user_prev.xlsx")
    write_master_xlsx(cycles, out)

    # 自动链不应覆盖用户已设置的非 None 值
    assert period_b.prev_recognition_rate == 0.5
    print("user-set prev preserved OK")


def test_master_xlsx_empty_raises():
    from agent.shenglong.excel_writer import write_master_xlsx
    import pytest
    with pytest.raises(ValueError):
        write_master_xlsx([], Path("downloads/shenglong/_unit_test/empty.xlsx"))


def test_master_prev_deduction_matches_visible_sheet_formula():
    """回归：下周期 F 列“上周期结果”必须直接等于上一周期 D 列展示口径。

    Sheet1 扣重符合率 D10 的公式是 B10/C2（符合车数 / 周期内有效检判车次）。
    旧 bug 用了 deduction_evaluable 做分母，导致 F 列和上一周期 D 列不一致。
    """
    from openpyxl import load_workbook
    from agent.shenglong.excel_writer import write_master_xlsx
    from agent.shenglong.models import PeriodSummary

    p1 = PeriodSummary(
        cycle_label="2026.4.30 至 2026.5.13",
        start_date="2026-04-30",
        end_date="2026-05-13",
        judgable_trucks=51,
        main_within_10pct_count=4,
        deduction_compliant_count=24,
        deduction_evaluable=54,  # 故意与 judgable 不同，用来复现旧 bug
    )
    p2 = PeriodSummary(
        cycle_label="2026.5.14 至 2026.5.20",
        start_date="2026-05-14",
        end_date="2026-05-20",
        judgable_trucks=43,
        main_within_10pct_count=9,
        deduction_compliant_count=16,
        deduction_evaluable=36,
    )

    out = Path("downloads/shenglong/_unit_test/master_prev_deduction_visible.xlsx")
    write_master_xlsx([([], p1), ([], p2)], out)

    wb = load_workbook(str(out), data_only=False)
    ws = wb["统计周期概括"]
    # 第 1 期 D10 可见公式 = 24 / 51
    assert ws.cell(row=10, column=4).value == "=IFERROR(B10/C2,0)"
    # 第 2 期 F24 必须摘抄第 1 期 D10 展示口径，而不是 24/54
    expected_prev = 24 / 51
    actual_prev = float(ws.cell(row=24, column=6).value)
    assert abs(actual_prev - expected_prev) < 1e-9
    assert abs(actual_prev - (24 / 54)) > 1e-3

    # 累计符合率也使用同一展示口径：sum(符合) / sum(有效车次)
    expected_cumulative = (24 + 16) / (51 + 43)
    actual_cumulative = float(ws.cell(row=23, column=8).value)
    assert abs(actual_cumulative - expected_cumulative) < 1e-9
    print("master prev deduction visible formula OK")


def test_master_xlsx_cumulative_rates_in_sheet1():
    """累计准确率/符合率：从首期累计到当期，每期 H 列应有数值"""
    from openpyxl import load_workbook
    from agent.shenglong.excel_writer import write_master_xlsx

    trucks_a = _build_sample_trucks()
    day_a = aggregate_daily("2026-04-14", trucks_a)
    period_a = aggregate_period([day_a], "2026-04-14", "2026-04-22")

    trucks_b = _build_sample_trucks()
    day_b = aggregate_daily("2026-04-23", trucks_b)
    period_b = aggregate_period([day_b], "2026-04-23", "2026-04-29")

    cycles = [([day_a], period_a), ([day_b], period_b)]
    out = Path("downloads/shenglong/_unit_test/master_cumulative.xlsx")
    write_master_xlsx(cycles, out)

    # ---- 累计字段已被 write_master_xlsx 注入 ----
    assert period_a.cumulative_recognition_rate is not None
    assert period_b.cumulative_recognition_rate is not None
    # 第 2 期累计 = (a.main_within + b.main_within) / (a.judgable + b.judgable)
    expected_b_cum = (
        (period_a.main_within_10pct_count + period_b.main_within_10pct_count)
        / (period_a.judgable_trucks + period_b.judgable_trucks)
    )
    assert abs(period_b.cumulative_recognition_rate - expected_b_cum) < 1e-9

    # ---- Sheet1 H 列：H4=标签, H5/H6 合并=数值 ----
    wb = load_workbook(str(out))
    ws = wb["统计周期概括"]
    assert ws.cell(row=4, column=8).value == "累计准确率"
    assert ws.cell(row=8, column=8).value == "累计符合率"
    # 第 1 期：H5（合并左上角）= 累计识别率（即第 1 期识别率本身，因为只有它）
    h5 = ws.cell(row=5, column=8).value
    assert h5 is not None and h5 != ""
    assert abs(float(h5) - period_a.cumulative_recognition_rate) < 1e-9
    # 第 2 期：H19（=15+4）= 累计识别率
    h19 = ws.cell(row=19, column=8).value
    assert abs(float(h19) - period_b.cumulative_recognition_rate) < 1e-9
    # 列宽 H 应为 14
    assert ws.column_dimensions["H"].width == 14
    print("master cumulative rates OK")


def test_master_xlsx_cumulative_sheet():
    """新增 Sheet「累计统计」：按截图式结构列出各期与 Tol 合计。"""
    from openpyxl import load_workbook
    from agent.shenglong.excel_writer import write_master_xlsx
    from agent.shenglong.models import PeriodSummary

    p1 = PeriodSummary(
        cycle_label="2026.4.14 至 2026.4.22",
        start_date="2026-04-14",
        end_date="2026-04-22",
        judgable_trucks=44,
        main_within_10pct_count=3,
        deduction_compliant_count=9,
    )
    p2 = PeriodSummary(
        cycle_label="2026.4.23 至 2026.4.29",
        start_date="2026-04-23",
        end_date="2026-04-29",
        judgable_trucks=17,
        main_within_10pct_count=1,
        deduction_compliant_count=5,
    )
    p3 = PeriodSummary(
        cycle_label="2026.4.30 至 2026.5.13",
        start_date="2026-04-30",
        end_date="2026-05-13",
        judgable_trucks=51,
        main_within_10pct_count=4,
        deduction_compliant_count=24,
    )
    p4 = PeriodSummary(
        cycle_label="2026.5.14 至 2026.5.20",
        start_date="2026-05-14",
        end_date="2026-05-20",
        judgable_trucks=43,
        main_within_10pct_count=9,
        deduction_compliant_count=16,
    )

    out = Path("downloads/shenglong/_unit_test/master_cumulative_sheet.xlsx")
    write_master_xlsx([([], p1), ([], p2), ([], p3), ([], p4)], out)

    wb = load_workbook(str(out), data_only=False)
    assert wb.sheetnames == ["统计周期概括", "累计统计", "检判统计详情"]
    ws = wb["累计统计"]
    assert ws.cell(row=1, column=1).value == "盛隆检判累计统计"
    assert ws.cell(row=2, column=3).value == "识别准确率"
    assert ws.cell(row=2, column=6).value == "扣重符合率"
    assert ws.cell(row=4, column=1).value == "第1期"
    assert ws.cell(row=4, column=3).value == 44
    assert ws.cell(row=4, column=4).value == 3
    assert ws.cell(row=4, column=5).value == "=IFERROR(D4/C4,0)"
    assert ws.cell(row=7, column=1).value == "第4期"
    assert ws.cell(row=7, column=6).value == 16

    # 第 8 行留白，第 9 行 Tol 合计，结构贴近用户截图。
    assert ws.cell(row=9, column=1).value == "Tol"
    assert ws.cell(row=9, column=3).value == "=SUM(C4:C7)"
    assert ws.cell(row=9, column=4).value == "=SUM(D4:D7)"
    assert ws.cell(row=9, column=5).value == "=IFERROR(D9/C9,0)"
    assert ws.cell(row=9, column=6).value == "=SUM(F4:F7)"
    assert ws.cell(row=9, column=7).value == "=IFERROR(F9/C9,0)"
    print("master cumulative sheet OK")


def test_master_export_ranges_group_chain():
    """工具入口：一个 cycle.ranges 里多个日期段才会合并成一个有效周期。

    模拟用户的 4 段输入，第 3 个 cycle 含 2 个 ranges，应产出 3 个有效周期。
    不真连 VPN，monkey-patch shenglong_client.build_range_stats 以构造的 stats 替代。
    """
    from agent.core import SteelCoilAgent
    from openpyxl import load_workbook

    # 准备 4 段对应的 fake stats（每段一个日期一辆车，便于查证日期合并）
    seg_dates = [
        ("2026-04-14", "2026-04-22", "2026-04-15"),  # 段 1，数据日期 4-15
        ("2026-04-23", "2026-04-29", "2026-04-24"),  # 段 2
        ("2026-04-30", "2026-05-06", "2026-05-01"),  # 第 3 期的第 1 段
        ("2026-05-07", "2026-05-13", "2026-05-08"),  # 段 4
    ]
    seg_to_day = {}
    for sd, ed, day in seg_dates:
        trucks = _build_sample_trucks()
        seg_to_day[(sd, ed)] = aggregate_daily(day, trucks[:1])

    def fake_build_range_stats(start_date: str, end_date: str):
        return [seg_to_day[(start_date, end_date)]]

    agent = SteelCoilAgent.__new__(SteelCoilAgent)  # 跳过 __init__
    # shenglong_client 是惰性 property → 直接预设私有属性绕过 client 实例化
    agent._shenglong_client = type("FakeClient", (), {
        "build_range_stats": staticmethod(fake_build_range_stats),
    })()

    cycles_input = [
        {"ranges": [{"start_date": "2026-04-14", "end_date": "2026-04-22"}]},
        {"ranges": [{"start_date": "2026-04-23", "end_date": "2026-04-29"}]},
        {"ranges": [
            {"start_date": "2026-04-30", "end_date": "2026-05-06"},
            {"start_date": "2026-05-07", "end_date": "2026-05-13"},
        ]},
    ]
    result = agent._tool_shenglong_export_master(cycles_input)

    assert "error" not in result, f"工具出错: {result}"
    assert result["cycle_count"] == 3, (
        f"期望 3 个有效周期（第 3、4 段合并），实际 {result['cycle_count']}"
    )

    # 第 3 期 cycle_label 应包含合并标记
    third_label = result["cycles"][2]["cycle_label"]
    assert "合并" in third_label, f"第 3 期应标记合并，实际: {third_label}"
    assert "2026-04-30" in third_label
    assert "2026-05-13" in third_label

    # 验证生成的 xlsx：Sheet2 应有 3 段（不是 4 段）
    wb = load_workbook(result["xlsx_path"])
    ws_d = wb["检判统计详情"]
    section_titles = [
        ws_d.cell(row=r, column=1).value
        for r in range(1, ws_d.max_row + 1)
        if isinstance(ws_d.cell(row=r, column=1).value, str)
        and ws_d.cell(row=r, column=1).value.startswith("第 ")
    ]
    assert len(section_titles) == 3, (
        f"Sheet2 应有 3 段，实际 {len(section_titles)}: {section_titles}"
    )
    assert "第 3 期" in section_titles[2]

    # 累计值也要在最后一期
    last_cycle = result["cycles"][-1]
    assert last_cycle["cumulative_recognition_rate"] is not None
    print("master ranges group OK")


def test_master_export_legacy_merge_with_next_ignored():
    """防回归：旧结构里的 merge_with_next 不再触发合并，避免 LLM 误传导致错误周期。

    用户输入 2026-04-30~2026-05-13、2026-05-14~2026-05-20 时，
    如果没有明确“当作一个统计周期”，这两段必须独立成两个周期。
    """
    from agent.core import SteelCoilAgent

    seg_dates = [
        ("2026-04-30", "2026-05-13", "2026-05-01"),
        ("2026-05-14", "2026-05-20", "2026-05-15"),
    ]
    seg_to_day = {}
    for sd, ed, day in seg_dates:
        trucks = _build_sample_trucks()
        seg_to_day[(sd, ed)] = aggregate_daily(day, trucks[:1])

    def fake_build_range_stats(start_date, end_date):
        return [seg_to_day[(start_date, end_date)]]

    agent = SteelCoilAgent.__new__(SteelCoilAgent)
    agent._shenglong_client = type("FakeClient", (), {
        "build_range_stats": staticmethod(fake_build_range_stats),
    })()

    cycles_input = [
        {
            "start_date": "2026-04-30",
            "end_date": "2026-05-13",
            "merge_with_next": True,  # 即便旧字段被误传，也必须忽略
        },
        {"start_date": "2026-05-14", "end_date": "2026-05-20"},
    ]
    result = agent._tool_shenglong_export_master(cycles_input)
    assert "error" not in result
    assert result["cycle_count"] == 2
    assert result["cycles"][0]["cycle_label"].startswith("2026.4.30 至 2026.5.13")
    assert result["cycles"][1]["cycle_label"].startswith("2026.5.14 至 2026.5.20")
    assert all("合并" not in c["cycle_label"] for c in result["cycles"])
    print("legacy merge_with_next ignored OK")


def test_master_export_no_merge_default_independent():
    """无 merge_with_next 字段 → 每段独立成周期（向后兼容）"""
    from agent.core import SteelCoilAgent

    seg_dates = [
        ("2026-04-14", "2026-04-22", "2026-04-15"),
        ("2026-04-23", "2026-04-29", "2026-04-24"),
    ]
    seg_to_day = {}
    for sd, ed, day in seg_dates:
        trucks = _build_sample_trucks()
        seg_to_day[(sd, ed)] = aggregate_daily(day, trucks[:1])

    def fake_build_range_stats(start_date, end_date):
        return [seg_to_day[(start_date, end_date)]]

    agent = SteelCoilAgent.__new__(SteelCoilAgent)
    agent._shenglong_client = type("FakeClient", (), {
        "build_range_stats": staticmethod(fake_build_range_stats),
    })()

    cycles_input = [
        {"start_date": "2026-04-14", "end_date": "2026-04-22"},
        {"start_date": "2026-04-23", "end_date": "2026-04-29"},
    ]
    result = agent._tool_shenglong_export_master(cycles_input)
    assert result["cycle_count"] == 2
    for c in result["cycles"]:
        assert "合并" not in c["cycle_label"]
    print("master no-merge default OK")


def test_heavy_normalized_formula_example():
    """用户截图公式：45/(45+35)=56.25%，35/(45+35)=43.75%。"""
    from agent.shenglong.calculator import _normalized_target_rates
    from agent.shenglong.dict import HEAVY_STEEL_TYPES

    mats = [
        MaterialRate(steel_type=1, rate=45.0),
        MaterialRate(steel_type=2, rate=35.0),
        MaterialRate(steel_type=13, rate=10.0),  # 厚剪：被剔出分母
        MaterialRate(steel_type=4, rate=10.0),   # 剪料1：被剔出分母
    ]
    rates = _normalized_target_rates(mats, HEAVY_STEEL_TYPES)
    assert abs(rates[1] - 56.25) < 1e-9
    assert abs(rates[2] - 43.75) < 1e-9
    assert rates[3] == 0.0
    print("heavy normalized formula OK")


def test_aggregate_period_heavy_normalized_metric():
    """重废归一化周期聚合：只改变识别率口径，扣重等统计保持原逻辑。"""
    # t1：人工包含非重废，归一化后与 AI 完全一致 → 正确
    t1 = TruckStat(
        date="2026-05-01",
        car_number="A1",
        station_number=1,
        flow_code="f1",
        manual_materials=[
            MaterialRate(1, 45.0),
            MaterialRate(2, 35.0),
            MaterialRate(13, 10.0),
            MaterialRate(4, 10.0),
        ],
        ai_materials=[
            MaterialRate(1, 56.25),
            MaterialRate(2, 43.75),
        ],
        deduction_compliant=True,
        price_diff=20.0,
    )
    # t2：双方都有重废，但归一化后主重废类不同 → 计入分母，不正确
    t2 = TruckStat(
        date="2026-05-01",
        car_number="A2",
        station_number=1,
        flow_code="f2",
        manual_materials=[
            MaterialRate(1, 45.0),
            MaterialRate(2, 35.0),
            MaterialRate(13, 20.0),
        ],
        ai_materials=[
            MaterialRate(1, 40.0),
            MaterialRate(2, 60.0),
        ],
        deduction_compliant=False,
        price_diff=80.0,
    )
    # t3：人工没有重废1/2/3 → 不进入新准确率分母，但扣重/价格仍正常统计
    t3 = TruckStat(
        date="2026-05-01",
        car_number="A3",
        station_number=1,
        flow_code="f3",
        manual_materials=[
            MaterialRate(13, 70.0),
            MaterialRate(4, 30.0),
        ],
        ai_materials=[
            MaterialRate(1, 100.0),
        ],
        deduction_compliant=True,
        price_diff=130.0,
    )

    day = aggregate_daily("2026-05-01", [t1, t2, t3])
    period = aggregate_period_heavy_normalized(
        [day], "2026-05-01", "2026-05-01"
    )

    assert period.recognition_section_title == "重废归一化识别率"
    assert period.judgable_trucks == 2
    assert period.main_name_match_count == 1
    assert period.main_within_10pct_count == 1
    assert abs(period.recognition_rate_pct - 50.0) < 1e-9
    # 扣重口径不变：3 辆均有 deduction_compliant
    assert period.deduction_evaluable == 3
    assert period.deduction_compliant_count == 2
    # 价格分桶仍保留原口径
    assert period.price_diff_lt30 == 1
    assert period.price_diff_50_100 == 1
    assert period.price_diff_gt100 == 1
    print("heavy normalized period OK")


def test_heavy_master_tool_generates_distinct_report():
    """新工具应生成独立的重废归一化主表，不覆盖普通主表。"""
    from agent.core import SteelCoilAgent
    from openpyxl import load_workbook

    trucks = [
        TruckStat(
            date="2026-05-01",
            car_number="A1",
            station_number=1,
            flow_code="f1",
            manual_materials=[
                MaterialRate(13, 40.0),  # 原始主料是厚剪
                MaterialRate(1, 30.0),
                MaterialRate(2, 20.0),
                MaterialRate(4, 10.0),
            ],
            ai_materials=[
                MaterialRate(1, 60.0),
                MaterialRate(2, 40.0),
            ],
            deduction_compliant=True,
        ),
        TruckStat(
            date="2026-05-01",
            car_number="A2",
            station_number=1,
            flow_code="f2",
            manual_materials=[
                MaterialRate(13, 70.0),  # 人工无任意重废1/2/3
                MaterialRate(4, 30.0),
            ],
            ai_materials=[
                MaterialRate(2, 100.0),
            ],
            deduction_compliant=True,
        )
    ]
    day = aggregate_daily("2026-05-01", trucks)

    def fake_build_range_stats(_start_date, _end_date):
        return [day]

    agent = SteelCoilAgent.__new__(SteelCoilAgent)
    agent._shenglong_client = type("FakeClient", (), {
        "build_range_stats": staticmethod(fake_build_range_stats),
    })()

    result = agent._tool_shenglong_export_master(
        [{"start_date": "2026-05-01", "end_date": "2026-05-07"}],
        heavy_normalized=True,
    )
    assert "error" not in result
    assert result["metric_label"] == "重废1/2/3归一化准确率"
    assert "_重废归一化_" in result["xlsx_path"]

    wb = load_workbook(result["xlsx_path"])
    ws = wb["统计周期概括"]
    assert ws.cell(row=3, column=2).value == "重废归一化识别率"
    assert ws.cell(row=5, column=4).value == "重废归一化准确率"
    assert ws.cell(row=6, column=3).value == 1

    # Sheet2 的主料型对比列也必须切换为归一化后的重废1/2/3视图：
    # A1 原始人工主料是“厚剪”，但重废归一化后应显示“重废1 60.00”
    ws_d = wb["检判统计详情"]
    assert ws_d.cell(row=7, column=2).value == "A1"
    assert ws_d.cell(row=7, column=4).value == "重废1"
    assert float(ws_d.cell(row=7, column=5).value) == 60.0
    assert ws_d.cell(row=7, column=6).value == "重废1"
    assert float(ws_d.cell(row=7, column=7).value) == 60.0

    # A2 人工没有重废1/2/3，不进入准确率统计；Sheet2 对比列显示不可判定
    assert ws_d.cell(row=8, column=2).value == "A2"
    assert ws_d.cell(row=8, column=4).value == "--"
    assert ws_d.cell(row=8, column=5).value == "/"
    assert ws_d.cell(row=8, column=8).value in (None, "")
    print("heavy master tool OK")


def test_excel_summary_sheet1_first_period_no_prev():
    """首期：未提供上周期 → F6/F10 留空，G5/G9 写 / """
    from openpyxl import load_workbook

    trucks = _build_sample_trucks()
    day1 = aggregate_daily("2026-04-22", trucks[:2])
    day2 = aggregate_daily("2026-04-23", trucks[2:])
    period = aggregate_period([day1, day2], "2026-04-22", "2026-04-23")
    # 不设 prev_*

    out = Path("downloads/shenglong/_unit_test/report_first_period.xlsx")
    write_stats_xlsx([day1, day2], out, period_summary=period)

    wb = load_workbook(str(out))
    ws = wb["统计周期概括"]
    # F6 留空，G5 = "/"
    assert ws.cell(row=6, column=6).value in (None, "")
    assert ws.cell(row=5, column=7).value == "/"
    assert ws.cell(row=10, column=6).value in (None, "")
    assert ws.cell(row=9, column=7).value == "/"
    print("sheet1 first period OK")


if __name__ == "__main__":
    test_dict()
    test_shenglong_record_station_number_list()
    test_operator_from_dict()
    test_normalize_deduction_kg_to_ton()
    test_operator_from_dict_kg_correction()
    test_calc_truck_with_kg_input()
    test_calc()
    test_aggregate_and_summary()
    test_excel()
    test_period_summary_aggregation()
    test_excel_with_summary_sheet()
    test_excluded_operators_filtered()
    test_all_operators_excluded_means_manual_missing()
    test_one_side_missing_not_counted()
    test_excel_summary_sheet1_with_prev_period()
    test_excel_summary_sheet1_first_period_no_prev()
    test_master_xlsx_two_cycles_with_auto_prev_chain()
    test_master_xlsx_user_provided_prev_not_overridden()
    test_master_xlsx_empty_raises()
    test_master_prev_deduction_matches_visible_sheet_formula()
    test_master_xlsx_cumulative_rates_in_sheet1()
    test_master_xlsx_cumulative_sheet()
    test_master_export_ranges_group_chain()
    test_master_export_legacy_merge_with_next_ignored()
    test_master_export_no_merge_default_independent()
    test_heavy_normalized_formula_example()
    test_aggregate_period_heavy_normalized_metric()
    test_heavy_master_tool_generates_distinct_report()
    print("\nAll shenglong unit smoke tests PASSED")
