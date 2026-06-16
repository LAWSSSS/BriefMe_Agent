"""用友检判统计模块单元/组件 smoke 测试（不需要访问真实 API / Playwright）

运行：
  python tests/test_yongyou_unit.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from agent.yongyou.config import create_settings, DEFAULT_DOWNLOAD_DIR
from agent.yongyou.page_actions import (
    ALLOWED_GRADES,
    TABLE_COLUMNS,
    is_allowed_grade,
    parse_primary_grade,
)
from agent.yongyou.downloader import get_image_filename, get_plate_from_filename
from agent.yongyou.excel_builder import (
    _find_hash,
    _format_primary_grade,
    _index_files_by_suffix,
    _scale_image,
    build_stat_excel,
)


# =====================================================================
# config
# =====================================================================
def test_create_settings():
    s = create_settings(username="testuser", password="testpass")
    assert s.username == "testuser"
    assert s.password == "testpass"
    assert s.base_url == "http://172.26.46.12:8890"
    assert s.record_url == f"{s.base_url}/imp-ib-iv-igs-fe/ibd/igs/scheme/#/record"
    assert s.download_dir.resolve() == DEFAULT_DOWNLOAD_DIR.resolve()
    assert s.cache_dir.exists()

    # 自定义下载目录
    custom = create_settings(
        username="u", password="p",
        download_dir="/tmp/test_yy_dl",
    )
    assert custom.download_dir == Path("/tmp/test_yy_dl").resolve()
    print("create_settings OK")


# =====================================================================
# page_actions 纯函数
# =====================================================================
def test_parse_primary_grade():
    # 标准格式
    assert parse_primary_grade("精炉料-等级二-80%,精炉料-等级一-20%") == "精炉料-等级二"
    assert parse_primary_grade("重废-等级一-100%") == "重废-等级一"
    assert parse_primary_grade("精炉料-等级三-50%,中废-50%") == "精炉料-等级三"

    # 无百分号
    assert parse_primary_grade("精炉料-等级二") == "精炉料-等级二"

    # 边界值
    assert parse_primary_grade("") == ""
    assert parse_primary_grade("无等级文字") == "无等级文字"
    print("parse_primary_grade OK")


def test_is_allowed_grade():
    assert is_allowed_grade("精炉料-等级二")
    assert is_allowed_grade("精炉料-等级三")
    assert is_allowed_grade("重废-等级一")
    assert is_allowed_grade("重废-等级二")
    assert not is_allowed_grade("精炉料-等级一")
    assert not is_allowed_grade("中废")
    assert not is_allowed_grade("")
    print("is_allowed_grade OK")


def test_allowed_grades_set():
    assert ALLOWED_GRADES == {"精炉料-等级二", "精炉料-等级三", "重废-等级一", "重废-等级二"}
    print("allowed_grades_set OK")


def test_table_columns():
    # 确保列定义无重复 key，去掉序号后第一列是 carNumber
    col_keys = [c[0] for c in TABLE_COLUMNS]
    assert col_keys[0] == "carNumber"
    assert "index" not in col_keys  # 已去掉序号
    assert len(col_keys) == len(set(col_keys))  # 无重复
    # 每个条目 3 元组
    for entry in TABLE_COLUMNS:
        assert len(entry) == 3
        assert isinstance(entry[0], str)
        assert isinstance(entry[1], str)
        assert isinstance(entry[2], int)
    print("table_columns OK")


# =====================================================================
# downloader 纯函数
# =====================================================================
def test_get_image_filename():
    assert get_image_filename("http://host/a/b/c/abc.jpg") == "abc.jpg"
    assert get_image_filename("/uploadFile/2026/06/14/foo.png") == "foo.png"
    # urlparse 把 "no/slash.jpg" 解析为 path="no/slash.jpg" → basename="slash.jpg"
    assert get_image_filename("no/slash.jpg") == "slash.jpg"
    print("get_image_filename OK")


def test_get_plate_from_filename():
    # 原图
    name1 = "GW9889-b24747b5998543548df17dfad281f58c-20260614144856_1.jpg"
    assert get_plate_from_filename(name1) == "GW9889"

    # visualize 前缀
    name2 = "visualize_GW9889-b24747b5998543548df17dfad281f58c-20260614144856_1.jpg"
    assert get_plate_from_filename(name2) == "GW9889"

    # visualize_superdetect_ 前缀
    name3 = "visualize_superdetect_GW9889-b24747b5998543548df17dfad281f58c-20260614144858_2.jpg"
    assert get_plate_from_filename(name3) == "GW9889"

    # visualize_special_ 前缀
    name4 = "visualize_special_GW9889-b24747b5998543548df17dfad281f58c-20260614144858_2.jpg"
    assert get_plate_from_filename(name4) == "GW9889"

    # 中文车牌
    name5 = "visualize_苏GW9889-b24747b5998543548df17dfad281f58c-20260614144856_1.jpg"
    assert get_plate_from_filename(name5) == "苏GW9889"

    # 未知格式兜底
    assert get_plate_from_filename("no-match-at-all") == "unknown"
    print("get_plate_from_filename OK")


# =====================================================================
# excel_builder 纯函数
# =====================================================================
def test_format_primary_grade():
    assert _format_primary_grade("精炉料-等级二-80%,精炉料-等级一-20%") == "精炉料2"
    assert _format_primary_grade("重废-等级一-100%") == "重废1"
    assert _format_primary_grade("精炉料-等级三-50%,中废-50%") == "精炉料3"
    assert _format_primary_grade("重废-等级二") == "重废2"
    assert _format_primary_grade("") == ""
    assert _format_primary_grade("不匹配的文本") == ""
    print("format_primary_grade OK")


def test_index_files_by_suffix(tmp_path: Path | None = None):
    # 用临时目录模拟
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        d = Path(td)
        # 创建一些文件
        (d / "pie_1.jpg").write_bytes(b"1")
        (d / "pie_2.jpg").write_bytes(b"2")
        (d / "pie_no_suffix.png").write_bytes(b"0")  # 无后缀 → index 0
        (d / "not_an_image.txt").write_bytes(b"x")
        (d / "pie_10.jpg").write_bytes(b"10")

        result = _index_files_by_suffix(d)
        assert result[1] == d / "pie_1.jpg"
        assert result[2] == d / "pie_2.jpg"
        assert result[0] == d / "pie_no_suffix.png"  # 无后缀记 0
        assert result[10] == d / "pie_10.jpg"
        assert "not_an_image.txt" not in str(result.values())
        print("index_files_by_suffix OK")


def test_scale_image():
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        from PIL import Image
        d = Path(td)
        # 创建 400x300 的测试图片
        img_path = d / "test.jpg"
        img = Image.new("RGB", (400, 300), color="red")
        img.save(img_path)

        # 目标 200x150 (等比缩小 50%)
        w, h = _scale_image(img_path, 200, 150)
        assert w == 200
        assert h == 150

        # 目标 100x100 (宽受限)
        w2, h2 = _scale_image(img_path, 100, 100)
        assert w2 == 100
        assert h2 == 75  # 保持 4:3

        # 目标 500x500 (等比放大填满)
        w3, h3 = _scale_image(img_path, 500, 500)
        assert w3 == 500
        assert h3 == 375  # 4:3 等比填满宽

        # 文件不存在
        w4, h4 = _scale_image(d / "nope.jpg", 200, 150)
        assert w4 == 200 and h4 == 150
        print("scale_image OK")


def test_find_hash():
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        d = Path(td) / "原图"
        d.mkdir()
        hash_val = "ba52f84a642e4b6dbae450457bbea065"
        (d / f"GW9889-{hash_val}-20260614144856_1.jpg").write_bytes(b"x")
        (d / "other_file_no_hash.jpg").write_bytes(b"y")

        assert _find_hash(d.parent) == hash_val

        # 空目录
        empty_dir = Path(td) / "empty"
        empty_dir.mkdir()
        assert _find_hash(empty_dir) == ""
        print("find_hash OK")


# =====================================================================
# build_stat_excel smoke（端到端）
# =====================================================================
def _make_sample_vehicle(tmp_root: Path, plate: str) -> tuple[str, dict, Path]:
    """创建一辆虚拟车的目录结构 + 表行数据"""
    vdir = tmp_root / plate
    for sub in ["饼图", "原图", "判级识别图", "夹杂物图", "危险物图"]:
        (vdir / sub).mkdir(parents=True)

    hash_val = "abc123def456abc123def456abc123de"
    # 饼图：无后缀 + _1 ~ _3
    from PIL import Image
    img = Image.new("RGB", (200, 150), color="blue")
    img.save(vdir / "饼图" / "废钢加工中心2工位.png")  # index 0
    img.save(vdir / "饼图" / "废钢加工中心2工位_1.png")  # index 1
    img.save(vdir / "饼图" / "废钢加工中心2工位_2.png")  # index 2
    img.save(vdir / "饼图" / "废钢加工中心2工位_3.png")  # index 3

    # 原图
    img.save(vdir / "原图" / f"{plate}-{hash_val}-20260614144856_1.jpg")
    img.save(vdir / "原图" / f"{plate}-{hash_val}-20260614144858_2.jpg")
    img.save(vdir / "原图" / f"{plate}-{hash_val}-20260614144901_3.jpg")
    img.save(vdir / "原图" / f"{plate}-{hash_val}-20260614144903_4.jpg")
    img.save(vdir / "原图" / f"{hash_val}-finish.jpg")  # 车底图

    # 判级识别图
    img.save(vdir / "判级识别图" / f"visualize_{plate}-{hash_val}-20260614144856_1.jpg")
    img.save(vdir / "判级识别图" / f"visualize_{plate}-{hash_val}-20260614144858_2.jpg")
    img.save(vdir / "判级识别图" / f"visualize_{plate}-{hash_val}-20260614144901_3.jpg")
    img.save(vdir / "判级识别图" / f"visualize_{plate}-{hash_val}-20260614144903_4.jpg")

    # 夹杂物图
    img.save(vdir / "夹杂物图" / f"visualize_special_{plate}-{hash_val}-20260614144858_2.jpg")
    img.save(vdir / "夹杂物图" / f"visualize_special_{plate}-{hash_val}-20260614144903_4.jpg")

    # 危险物图
    img.save(vdir / "危险物图" / f"visualize_superdetect_{plate}-{hash_val}-20260614144858_2.jpg")

    row_data = {
        "carNumber": plate,
        "ibGradingDef": "重废-等级一-100%",
        "afGradingDef": "重废-等级一-100%,中废-0%",
        "ibOnRmimp": "120.0",
        "afRmimp": "115.0",
        "synthesizeThickness": "8.5",
        "ib2Price": "2800",
        "price": "2780",
        "punMount": "-20",
        "oversizeRatio": "5.2",
        "materialTypeMax": "重废-等级一",
        "showThickThin": "HM 60%",
        "sysSlagRatio": "3.1",
        "sysWarnSlagSum": "15.0",
        "sysWarnFee": "30.0",
        "baseRmimp": "80.0",
        "lowGradeRmimp": "10.0",
        "slagRmimp": "15.0",
        "compositeRmimp": "10.0",
        "rasWeatherRmimp": "0",
        "counterweightRmimp": "0",
        "pigIronRmimp": "0",
        "oilDegree": "轻微",
        "rustDegree": "无",
        "soilDegree": "轻微",
        "compositeDegree": "轻微",
        "suckerCount": "2%",
        "oilWarning": "0%",
        "sysMohuWarning": "1%",
    }
    return (plate, row_data, vdir)


def test_build_stat_excel():
    with TemporaryDirectory() as td:
        tmp = Path(td)
        v = _make_sample_vehicle(tmp, "苏GW9889")
        vehicles = [v]
        out = tmp / "用友检判统计_20260614.xlsx"

        build_stat_excel(vehicles, out)
        assert out.exists()
        assert out.stat().st_size > 5000

        # 验证 sheet 名
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "苏GW9889_重废1" in wb.sheetnames
        ws = wb["苏GW9889_重废1"]

        # 表头
        assert ws.cell(row=1, column=1).value == "车牌号"
        # 数据
        assert ws.cell(row=2, column=1).value == "苏GW9889"

        # 列宽统一 21
        from openpyxl.utils import get_column_letter
        assert ws.column_dimensions[get_column_letter(1)].width == 21

        # 冻结
        assert ws.freeze_panes == "A3"

        # 有合并单元格（饼图）
        assert len(ws.merged_cells.ranges) >= 1

        print("build_stat_excel OK")


if __name__ == "__main__":
    test_create_settings()
    test_parse_primary_grade()
    test_is_allowed_grade()
    test_allowed_grades_set()
    test_table_columns()
    test_get_image_filename()
    test_get_plate_from_filename()
    test_format_primary_grade()
    test_index_files_by_suffix()
    test_scale_image()
    test_find_hash()
    test_build_stat_excel()
    print("\nAll yongyou unit smoke tests PASSED")
